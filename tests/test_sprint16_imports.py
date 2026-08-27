import io
from datetime import date
from decimal import Decimal

import pytest
from django.core.exceptions import ValidationError
from django.core.files.uploadedfile import SimpleUploadedFile
from django.urls import reverse
from openpyxl import load_workbook

from apps.imports.services import (
    build_template_workbook,
    cancel_import_batch,
    confirm_import_batch,
    upload_and_validate_import,
)
from apps.supplies.models import (
    SupplyCustody,
    SupplyCustodyMovement,
    SupplyStockBalance,
    SupplyStockLedger,
)
from tests.test_sprint15_support import (
    make_company,
    make_department,
    make_employee,
    make_supply_category,
    make_supply_item,
    make_user,
)


pytestmark = pytest.mark.django_db


@pytest.fixture(autouse=True)
def isolated_import_media(settings, tmp_path):
    settings.MEDIA_ROOT = tmp_path / "media"


def opening_custody_workbook(company, rows):
    data = build_template_workbook("opening_custody", company=company)
    workbook = load_workbook(io.BytesIO(data))
    sheet = workbook["期初保管导入"]
    for row in rows:
        sheet.append(row)
    output = io.BytesIO()
    workbook.save(output)
    workbook.close()
    return SimpleUploadedFile(
        "opening-custody.xlsx",
        output.getvalue(),
        content_type=(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ),
    )


def test_opening_custody_template_and_confirmation_are_atomic_idempotent_and_stock_neutral():
    company = make_company()
    actor = make_user("s16-opening-equipment", "equipment")
    department = make_department(company, "USE")
    employee = make_employee(company, department, "EMP")
    category = make_supply_category(company)
    durable = make_supply_item(
        company,
        category,
        "CHAIR",
        item_type="durable_quantity",
        unit="把",
    )
    template = load_workbook(io.BytesIO(build_template_workbook("opening_custody", company=company)))
    assert [cell.value for cell in template["期初保管导入"][1]] == [
        "物品编码",
        "责任部门编码",
        "责任员工编号",
        "数量",
        "单位成本",
        "开始日期",
        "备注",
    ]
    assert len(template["期初保管导入"].data_validations.dataValidation) == 2
    template.close()

    batch = upload_and_validate_import(
        actor=actor,
        company=company,
        import_type="opening_custody",
        uploaded_file=opening_custody_workbook(
            company,
            [
                [
                    durable.item_code,
                    department.code,
                    employee.employee_no,
                    2,
                    80,
                    date(2026, 8, 1),
                    "历史在管",
                ]
            ],
        ),
        idempotency_key="s16-opening-confirm",
    )
    assert batch.status == "validated"
    confirm_import_batch(actor=actor, batch=batch)
    batch.refresh_from_db()
    custody = SupplyCustody.objects.get(origin_import_row__batch=batch)
    movement = SupplyCustodyMovement.objects.get(to_custody=custody)
    assert batch.status == "confirmed"
    assert custody.parent_custody_id is None
    assert custody.origin_issue_line_id is None
    assert custody.current_quantity == Decimal("2.0000")
    assert custody.current_amount == Decimal("160.00")
    assert movement.action == "opening"
    assert movement.amount == Decimal("160.00")
    assert batch.rows.get().created_object_type == "SupplyCustody"
    assert not SupplyStockBalance.objects.exists()
    assert not SupplyStockLedger.objects.exists()
    confirm_import_batch(actor=actor, batch=batch)
    assert SupplyCustody.objects.count() == 1
    assert SupplyCustodyMovement.objects.count() == 1
    with pytest.raises(ValidationError, match="确认后不得破坏性回滚"):
        cancel_import_batch(actor=actor, batch=batch, reason="错误回滚")


def test_opening_custody_validation_rejects_consumable_employee_department_and_rolls_back_all():
    company = make_company()
    actor = make_user("s16-opening-invalid", "warehouse")
    department = make_department(company, "A")
    other_department = make_department(company, "B")
    employee = make_employee(company, other_department, "EMP-B")
    category = make_supply_category(company)
    consumable = make_supply_item(company, category, "PAPER")
    batch = upload_and_validate_import(
        actor=actor,
        company=company,
        import_type="opening_custody",
        uploaded_file=opening_custody_workbook(
            company,
            [
                [
                    consumable.item_code,
                    department.code,
                    employee.employee_no,
                    0,
                    -1,
                    "bad-date",
                    "错误行",
                ]
            ],
        ),
        idempotency_key="s16-opening-invalid",
    )
    assert batch.status == "invalid"
    fields = {error["field"] for error in batch.rows.get().errors_json}
    assert fields >= {"物品编码", "责任员工编号", "数量", "单位成本", "开始日期"}
    assert not SupplyCustody.objects.exists()
    assert not SupplyCustodyMovement.objects.exists()


def test_opening_custody_can_cancel_before_confirmation_but_not_confirm_after_cancel():
    company = make_company()
    actor = make_user("s16-opening-cancel", "finance")
    department = make_department(company)
    category = make_supply_category(company)
    durable = make_supply_item(
        company, category, item_type="durable_quantity"
    )
    batch = upload_and_validate_import(
        actor=actor,
        company=company,
        import_type="opening_custody",
        uploaded_file=opening_custody_workbook(
            company,
            [[durable.item_code, department.code, "", 1, 0, date(2026, 8, 1), "取消"]],
        ),
        idempotency_key="s16-opening-cancel",
    )
    cancel_import_batch(actor=actor, batch=batch, reason="不再导入")
    batch.refresh_from_db()
    assert batch.status == "cancelled"
    with pytest.raises(ValidationError, match="只能确认"):
        confirm_import_batch(actor=actor, batch=batch)
    assert not SupplyCustody.objects.exists()


def test_opening_custody_http_form_is_registered_and_role_scoped(client):
    company = make_company()
    equipment = make_user("s16-opening-http-equipment", "equipment")
    department = make_department(company)
    category = make_supply_category(company)
    durable = make_supply_item(
        company, category, item_type="durable_quantity"
    )
    client.force_login(equipment)
    assert client.get(
        reverse("imports:template", args=["opening_custody"])
    ).status_code == 200
    assert client.get(
        reverse("imports:upload", args=["opening_custody"])
    ).status_code == 200
    response = client.post(
        reverse("imports:upload", args=["opening_custody"]),
        {
            "import_type": "opening_custody",
            "file": opening_custody_workbook(
                company,
                [
                    [
                        durable.item_code,
                        department.code,
                        "",
                        1,
                        0,
                        date(2026, 8, 1),
                        "HTTP",
                    ]
                ],
            ),
        },
    )
    assert response.status_code == 302
    management = make_user("s16-opening-http-management", "management")
    client.force_login(management)
    assert client.get(
        reverse("imports:upload", args=["opening_custody"])
    ).status_code == 403
