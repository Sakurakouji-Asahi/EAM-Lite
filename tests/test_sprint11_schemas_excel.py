from __future__ import annotations

import io
import time
import zipfile
from datetime import date, datetime
from decimal import Decimal
from types import MappingProxyType

import pytest
from django.utils import timezone
from openpyxl import load_workbook

from apps.reports.excel import write_report_workbook, write_tplus_workbook
from apps.reports.queries import ReportDataset, TplusDataset
from apps.reports.schemas import (
    CellKind,
    REPORT_REGISTRY,
    REPORT_SCHEMA_VERSION,
    TOTAL_METRIC_REGISTRY,
    TPLUS_ASSET_COLUMNS,
    TPLUS_ENTRY_COLUMNS,
    TPLUS_SCHEMA_VERSION,
    TPLUS_TOTAL_METRICS,
    validate_totals,
)


EXPECTED_REPORT_TYPES = {
    "asset_ledger",
    "fixed_asset_detail",
    "depreciation_schedule",
    "depreciation_detail",
    "monthly_depreciation",
    "department_assets",
    "employee_assets",
    "equipment_list",
    "mold_tool_inspection_list",
    "inventory_results",
    "inventory_differences",
    "maintenance_plans",
    "maintenance_due",
    "maintenance_records",
    "offboarding_unresolved",
    "disposal_list",
    "tplus_reconciliation",
}


def _row_for(columns):
    values = {}
    for column in columns:
        if column.kind in {CellKind.MONEY, CellKind.DECIMAL}:
            values[column.key] = Decimal("12.34")
        elif column.kind == CellKind.RATE:
            values[column.key] = Decimal("0.05")
        elif column.kind == CellKind.INTEGER:
            values[column.key] = 7
        elif column.kind == CellKind.DATE:
            values[column.key] = date(2026, 8, 13)
        elif column.kind == CellKind.DATETIME:
            values[column.key] = timezone.now()
        elif column.kind == CellKind.IDENTIFIER:
            values[column.key] = "00123"
        else:
            values[column.key] = "=HYPERLINK(\"https://invalid.example\")"
    return MappingProxyType(values)


def _tplus_dataset():
    definition = REPORT_REGISTRY["tplus_reconciliation"]
    asset_row = dict(_row_for(definition.columns))
    asset_row.update(
        {
            "asset_code": "00123",
            "tplus_card_code": "0000456",
            "asset_name": "=DDE|danger",
            "opening_accumulated_depreciation": Decimal("10.00"),
            "automatic_depreciation": Decimal("2.00"),
            "manual_depreciation": Decimal("3.00"),
            "adjustment_net": Decimal("-1.00"),
            "reversal_net": Decimal("1.00"),
            "ending_accumulated_depreciation": Decimal("15.00"),
            "original_cost": Decimal("100.00"),
            "impairment": Decimal("5.00"),
            "ending_book_value": Decimal("80.00"),
            "disposal_income": Decimal("0.00"),
        }
    )
    entry_row = dict(_row_for(TPLUS_ENTRY_COLUMNS))
    entry_row.update(
        {
            "asset_code": "00123",
            "tplus_card_code": "0000456",
            "amount": Decimal("-2.00"),
        }
    )
    totals = {
        key: Decimal(str(asset_row[key])).quantize(Decimal("0.01"))
        for key in TPLUS_TOTAL_METRICS
    }
    return TplusDataset(
        definition=definition,
        asset_rows=(MappingProxyType(asset_row),),
        entry_rows=(MappingProxyType(entry_row),),
        filters=MappingProxyType({"period": "2026-08"}),
        data_snapshot_at=timezone.now(),
        totals=MappingProxyType(totals),
    )


def test_registry_contains_exactly_seventeen_approved_report_types_and_two_versions():
    assert set(REPORT_REGISTRY) == EXPECTED_REPORT_TYPES
    assert all(item.key == key for key, item in REPORT_REGISTRY.items())
    assert {
        item.schema_version for item in REPORT_REGISTRY.values() if not item.tplus
    } == {REPORT_SCHEMA_VERSION}
    assert REPORT_REGISTRY["tplus_reconciliation"].schema_version == TPLUS_SCHEMA_VERSION
    assert len(TPLUS_ASSET_COLUMNS) == 30
    assert len({column.key for column in TPLUS_ASSET_COLUMNS}) == 30
    assert tuple(REPORT_REGISTRY["tplus_reconciliation"].total_metrics) == tuple(
        TPLUS_TOTAL_METRICS
    )
    assert TOTAL_METRIC_REGISTRY == {
        REPORT_SCHEMA_VERSION: frozenset(),
        TPLUS_SCHEMA_VERSION: frozenset(TPLUS_TOTAL_METRICS),
    }


def test_total_registry_requires_decimal_exact_set_and_generic_reports_have_no_totals():
    totals = {key: Decimal("0.00") for key in TPLUS_TOTAL_METRICS}
    validate_totals("tplus_reconciliation", TPLUS_SCHEMA_VERSION, totals)
    validate_totals("asset_ledger", REPORT_SCHEMA_VERSION, {})
    with pytest.raises(ValueError):
        validate_totals(
            "tplus_reconciliation",
            TPLUS_SCHEMA_VERSION,
            {key: value for key, value in totals.items() if key != "disposal_income"},
        )
    with pytest.raises(ValueError):
        validate_totals(
            "tplus_reconciliation",
            TPLUS_SCHEMA_VERSION,
            {**totals, "unapproved_total": Decimal("0.00")},
        )
    with pytest.raises(TypeError):
        validate_totals(
            "tplus_reconciliation",
            TPLUS_SCHEMA_VERSION,
            {**totals, "original_cost": 0},
        )


def test_generic_workbook_preserves_numeric_date_identifier_and_escapes_text():
    definition = REPORT_REGISTRY["fixed_asset_detail"]
    dataset = ReportDataset(
        definition=definition,
        rows=(_row_for(definition.columns),),
        filters=MappingProxyType(
            {
                "as_of_date": date(2026, 8, 13),
                "include_disposed": False,
                "minimum_cost": Decimal("12.34"),
            }
        ),
        data_snapshot_at=timezone.now(),
    )
    output = io.BytesIO()
    write_report_workbook(dataset, output, generated_at=timezone.now())
    output.seek(0)
    workbook = load_workbook(output, data_only=False)
    information = workbook["导出说明"]
    metadata = {row[0].value: row[1] for row in information.iter_rows()}
    assert metadata["生成完成时间"].data_type == "d"
    assert metadata["数据截止时间"].data_type == "d"
    assert metadata["Schema版本"].value == REPORT_SCHEMA_VERSION
    assert metadata["筛选条件：as_of_date"].data_type == "d"
    assert metadata["筛选条件：as_of_date"].value.date() == date(2026, 8, 13)
    assert metadata["筛选条件：include_disposed"].value is False
    assert metadata["筛选条件：include_disposed"].data_type == "b"
    assert metadata["筛选条件：minimum_cost"].data_type == "n"
    assert Decimal(str(metadata["筛选条件：minimum_cost"].value)) == Decimal("12.34")
    sheet = workbook[definition.sheet_name]
    cells = {
        column.key: sheet.cell(row=2, column=index)
        for index, column in enumerate(definition.columns, start=1)
    }
    assert cells["asset_code"].value == "00123"
    assert cells["asset_code"].number_format == "@"
    assert cells["original_cost"].data_type == "n"
    assert Decimal(str(cells["original_cost"].value)) == Decimal("12.34")
    assert cells["acquisition_date"].data_type == "d"
    assert cells["asset_name"].value.startswith("'=")
    assert cells["asset_name"].data_type != "f"


def test_tplus_workbook_has_exact_five_sheets_thirty_columns_and_safe_types():
    dataset = _tplus_dataset()
    output = io.BytesIO()
    write_tplus_workbook(
        dataset,
        output,
        export_id="00000000-0000-0000-0000-000000000001",
        company_name="测试公司",
        requested_by="finance",
        generated_at=timezone.now(),
    )
    output.seek(0)
    workbook = load_workbook(output, data_only=False)
    assert workbook.sheetnames == [
        "导出说明",
        "EAM固定资产明细",
        "本期折旧分录",
        "T+数据粘贴区",
        "对账差异",
    ]
    detail = workbook["EAM固定资产明细"]
    assert detail.max_column == 30
    assert tuple(cell.value for cell in detail[1]) == tuple(
        column.label for column in TPLUS_ASSET_COLUMNS
    )
    detail_cells = {
        column.key: detail.cell(row=2, column=index)
        for index, column in enumerate(TPLUS_ASSET_COLUMNS, start=1)
    }
    assert detail_cells["asset_code"].value == "00123"
    assert detail_cells["asset_code"].number_format == "@"
    assert detail_cells["tplus_card_code"].value == "0000456"
    assert detail_cells["original_cost"].data_type == "n"
    assert detail_cells["capitalization_date"].data_type == "d"
    assert detail_cells["asset_name"].value.startswith("'=")
    assert workbook["导出说明"].protection.sheet is True
    assert detail.protection.sheet is True
    assert workbook["本期折旧分录"].protection.sheet is True
    paste = workbook["T+数据粘贴区"]
    assert tuple(cell.value for cell in paste[1]) == (
        "T+资产卡片编码",
        "EAM资产编号（如T+已维护）",
        "资产名称",
        "原值",
        "期初累计折旧",
        "本期折旧",
        "期末累计折旧",
        "减值准备",
        "期末账面净值",
    )
    assert paste.max_column == 9
    assert paste["A2"].number_format == "@"
    assert paste["B2"].number_format == "@"
    difference = workbook["对账差异"]
    assert difference["A2"].data_type == "f"
    assert "'T+数据粘贴区'!$A$2:$A$5001" in difference["A2"].value
    assert "'T+数据粘贴区'!$B$2:$B$5001" in difference["A2"].value
    assert '"duplicate key"' in difference["A2"].value
    assert difference["F2"].data_type == "f"
    assert "'T+数据粘贴区'!$D$2:$D$5001" in difference["F2"].value
    assert difference["G2"].value == '=IF(A2="matched",E2-F2,"")'
    extra_status = difference["A3"].value
    assert "'T+数据粘贴区'!$A$2:$A$5001" in extra_status
    assert "'T+数据粘贴区'!$B$2:$B$5001" in extra_status
    assert '"duplicate key"' in extra_status


def test_tplus_formulas_prioritize_card_then_fallback_to_code_within_120_seconds():
    output = io.BytesIO()
    started = time.perf_counter()
    write_tplus_workbook(
        _tplus_dataset(),
        output,
        export_id="00000000-0000-0000-0000-000000000002",
        company_name="测试公司",
        requested_by="finance",
        generated_at=timezone.now(),
    )
    elapsed = time.perf_counter() - started
    assert elapsed <= 120
    assert output.tell() > 0

    output.seek(0)
    workbook = load_workbook(output, data_only=False)
    paste = workbook["T+数据粘贴区"]
    difference = workbook["对账差异"]
    assert paste.max_row == 5001
    assert difference.max_row == 5002

    card_range = "'T+数据粘贴区'!$A$2:$A$5001"
    code_range = "'T+数据粘贴区'!$B$2:$B$5001"
    status_formula = difference["A2"].value
    assert (
        f'IF(COUNTIF({card_range},C2)=1,"matched",'
        f'IF(COUNTIF({code_range},B2)>1,"duplicate key",'
        f'IF(COUNTIF({code_range},B2)=1,"matched","EAM only")))'
        in status_formula
    )
    lookup_formula = difference["F2"].value
    assert (
        f'IF(COUNTIF({card_range},C2)=1,'
        f'INDEX(\'T+数据粘贴区\'!$D$2:$D$5001,'
        f'MATCH(C2,{card_range},0)),'
        f'IF(COUNTIF({code_range},B2)=1,'
        in lookup_formula
    )

    extra_status = difference["A3"].value
    eam_card_range = "'EAM固定资产明细'!$B$2:$B$1048576"
    eam_code_range = "'EAM固定资产明细'!$A$2:$A$1048576"
    assert extra_status.index(eam_card_range) < extra_status.index(eam_code_range)
    assert extra_status.count(f"COUNTIF({eam_code_range},'T+数据粘贴区'!B2)") >= 2
    assert '"T+ only"' in extra_status


def test_tplus_rows_and_registered_totals_satisfy_both_decimal_identities():
    dataset = _tplus_dataset()
    row = dataset.asset_rows[0]
    assert row["ending_accumulated_depreciation"] == sum(
        (
            row["opening_accumulated_depreciation"],
            row["automatic_depreciation"],
            row["manual_depreciation"],
            row["adjustment_net"],
            row["reversal_net"],
        ),
        Decimal("0.00"),
    )
    assert row["ending_book_value"] == (
        row["original_cost"]
        - row["ending_accumulated_depreciation"]
        - row["impairment"]
    )
    assert dataset.totals == {
        key: sum((item[key] for item in dataset.asset_rows), Decimal("0.00"))
        for key in TPLUS_TOTAL_METRICS
    }


def test_tplus_xlsx_package_has_no_macros_or_external_links():
    output = io.BytesIO()
    write_tplus_workbook(
        _tplus_dataset(),
        output,
        export_id="00000000-0000-0000-0000-000000000001",
        company_name="测试公司",
        requested_by="finance",
        generated_at=datetime.now(tz=timezone.get_current_timezone()),
    )
    with zipfile.ZipFile(io.BytesIO(output.getvalue())) as archive:
        names = set(archive.namelist())
        assert not any(name.lower().endswith("vbaproject.bin") for name in names)
        assert not any("externalLinks" in name for name in names)
        assert "xl/connections.xml" not in names
        content_types = archive.read("[Content_Types].xml").decode("utf-8")
        assert "macroEnabled" not in content_types
        assert "vbaProject" not in content_types
