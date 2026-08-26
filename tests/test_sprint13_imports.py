import io

import pytest
from django.core.exceptions import PermissionDenied, ValidationError
from django.core.files.uploadedfile import SimpleUploadedFile
from django.urls import reverse
from openpyxl import load_workbook

from apps.audit.models import AuditLog
from apps.imports.services import (
    build_template_workbook,
    confirm_import_batch,
    upload_and_validate_import,
)
from apps.supplies.models import SupplyItem
from tests.test_sprint13_support import (
    make_company,
    make_supply_category,
    make_supply_item,
    make_supply_warehouse,
    make_user,
)


pytestmark = pytest.mark.django_db


@pytest.fixture(autouse=True)
def isolated_supply_import_media(settings, tmp_path):
    settings.MEDIA_ROOT = tmp_path / "media"


def item_workbook(company, rows):
    data = build_template_workbook("item_master", company=company)
    workbook = load_workbook(io.BytesIO(data))
    sheet = workbook["低值物品档案导入"]
    for row in rows:
        sheet.append(row)
    output = io.BytesIO()
    workbook.save(output)
    workbook.close()
    return SimpleUploadedFile(
        "item-master.xlsx",
        output.getvalue(),
        content_type=(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ),
    )


def test_item_master_template_has_fixed_headers_validations_and_numeric_example():
    company = make_company()
    data = build_template_workbook("item_master", company=company)
    workbook = load_workbook(io.BytesIO(data))
    sheet = workbook["低值物品档案导入"]
    assert [cell.value for cell in sheet[1]] == [
        "物品编码",
        "物品名称",
        "分类编码",
        "管理模式",
        "单位",
        "规格",
        "型号",
        "品牌",
        "最低库存",
        "默认仓库编码",
        "备注",
    ]
    assert any(
        validation.formula1 == '"consumable,durable_quantity"'
        for validation in sheet.data_validations.dataValidation
    )
    assert workbook["示例"]["I2"].value == 5
    assert isinstance(workbook["示例"]["I2"].value, int)
    workbook.close()


def test_item_import_reports_row_errors_and_writes_no_items():
    company = make_company()
    actor = make_user("s13-import-warehouse", "warehouse")
    make_supply_category(company, "OFFICE")
    make_supply_warehouse(company, "OFFICE-WH")
    upload = item_workbook(
        company,
        [
            ["PAPER", "复印纸", "OFFICE", "consumable", "箱", "", "", "", 5, "OFFICE-WH", ""],
            [" paper ", "重复编码", "OFFICE", "consumable", "箱", "", "", "", 0, "", ""],
            ["BAD-CAT", "分类错误", "MISSING", "consumable", "个", "", "", "", 0, "", ""],
            ["BAD-TYPE", "类型错误", "OFFICE", "serialized", "个", "", "", "", 0, "", ""],
            ["BAD-MIN", "负库存", "OFFICE", "consumable", "个", "", "", "", -1, "", ""],
            ["BAD-UNIT", "空单位", "OFFICE", "consumable", "", "", "", "", 0, "", ""],
        ],
    )
    batch = upload_and_validate_import(
        actor=actor,
        company=company,
        import_type="item_master",
        uploaded_file=upload,
        idempotency_key="s13-item-invalid",
    )
    assert batch.status == "invalid"
    assert batch.error_rows == 5
    assert not SupplyItem.objects.filter(company=company).exists()
    errors = [error for row in batch.rows.all() for error in row.errors_json]
    assert all({"field", "value", "reason"} <= set(error) for error in errors)
    assert {error["field"] for error in errors} >= {
        "物品编码",
        "分类编码",
        "管理模式",
        "最低库存",
        "单位",
    }


def test_item_import_confirmation_is_atomic_audited_and_idempotent():
    company = make_company()
    actor = make_user("s13-import-finance", "finance")
    make_supply_category(company, "OFFICE")
    make_supply_warehouse(company, "OFFICE-WH")
    batch = upload_and_validate_import(
        actor=actor,
        company=company,
        import_type="item_master",
        uploaded_file=item_workbook(
            company,
            [
                ["PAPER-A4", "A4 复印纸", "OFFICE", "consumable", "箱", "A4", "", "", 5, "OFFICE-WH", ""],
                ["CHAIR-01", "普通办公椅", "OFFICE", "durable_quantity", "把", "", "", "", 0, "", ""],
            ],
        ),
        idempotency_key="s13-item-valid",
    )
    assert batch.status == "validated"
    confirm_import_batch(actor=actor, batch=batch)
    batch.refresh_from_db()
    assert batch.status == "confirmed"
    assert batch.rows.filter(validation_status="created").count() == 2
    assert set(
        SupplyItem.objects.filter(company=company).values_list(
            "item_code", "item_type"
        )
    ) == {
        ("PAPER-A4", "consumable"),
        ("CHAIR-01", "durable_quantity"),
    }
    assert AuditLog.objects.filter(
        company=company,
        action="import_confirm",
        object_id=str(batch.pk),
    ).exists()
    assert AuditLog.objects.filter(
        company=company, action="supply_item_create"
    ).count() == 2

    confirm_import_batch(actor=actor, batch=batch)
    assert SupplyItem.objects.filter(company=company).count() == 2


def test_item_import_confirmation_conflict_rolls_back_entire_batch():
    company = make_company()
    actor = make_user("s13-import-admin", "system_admin")
    category = make_supply_category(company, "OFFICE")
    batch = upload_and_validate_import(
        actor=actor,
        company=company,
        import_type="item_master",
        uploaded_file=item_workbook(
            company,
            [
                ["RACE-A", "并发 A", "OFFICE", "consumable", "个", "", "", "", 0, "", ""],
                ["RACE-B", "并发 B", "OFFICE", "consumable", "个", "", "", "", 0, "", ""],
            ],
        ),
        idempotency_key="s13-item-race",
    )
    assert batch.status == "validated"
    make_supply_item(company, category, "RACE-A")

    with pytest.raises(ValidationError):
        confirm_import_batch(actor=actor, batch=batch)

    batch.refresh_from_db()
    assert batch.status == "validated"
    assert not SupplyItem.objects.filter(company=company, item_code="RACE-B").exists()
    assert not batch.rows.filter(validation_status="created").exists()
    assert AuditLog.objects.filter(
        company=company,
        action="import_confirm_failed",
        object_id=str(batch.pk),
    ).exists()


def test_equipment_import_may_only_include_durable_quantity_items():
    company = make_company()
    category = make_supply_category(company, "TOOLS")
    equipment = make_user("s13-import-equipment", "equipment")
    batch = upload_and_validate_import(
        actor=equipment,
        company=company,
        import_type="item_master",
        uploaded_file=item_workbook(
            company,
            [
                ["PAPER", "复印纸", "TOOLS", "consumable", "箱", "", "", "", 0, "", ""],
                ["HAMMER", "锤子", "TOOLS", "durable_quantity", "把", "", "", "", 0, "", ""],
            ],
        ),
        idempotency_key="s13-equipment-scope",
    )
    assert batch.status == "invalid"
    assert not SupplyItem.objects.filter(company=company).exists()

    employee = make_user("s13-import-employee", "employee")
    with pytest.raises(PermissionDenied):
        upload_and_validate_import(
            actor=employee,
            company=company,
            import_type="item_master",
            uploaded_file=item_workbook(company, []),
            idempotency_key="s13-employee-denied",
        )
    assert category.is_active


def test_item_import_http_endpoints_enforce_backend_permission(client):
    make_company()
    equipment = make_user("s13-import-http-equipment", "equipment")
    client.force_login(equipment)
    assert client.get(reverse("imports:template", args=["item_master"])).status_code == 200
    assert client.get(reverse("imports:upload", args=["item_master"])).status_code == 200

    management = make_user("s13-import-http-management", "management")
    client.force_login(management)
    assert client.get(reverse("imports:template", args=["item_master"])).status_code == 403
    assert client.get(reverse("imports:upload", args=["item_master"])).status_code == 403
