import io

import pytest
from django.core.exceptions import PermissionDenied, ValidationError
from django.core.files.uploadedfile import SimpleUploadedFile
from django.urls import reverse
from openpyxl import load_workbook

from apps.imports.services import (
    build_template_workbook,
    confirm_import_batch,
    upload_and_validate_import,
)
from apps.supplies.models import SupplyDocument, SupplyStockBalance, SupplyStockLedger
from tests.test_sprint14_support import (
    make_company,
    make_supply_category,
    make_supply_item,
    make_supply_warehouse,
    make_user,
)


pytestmark = pytest.mark.django_db


@pytest.fixture(autouse=True)
def isolated_import_media(settings, tmp_path):
    settings.MEDIA_ROOT = tmp_path / "media"


def opening_workbook(company, rows):
    data = build_template_workbook("opening_stock", company=company)
    workbook = load_workbook(io.BytesIO(data))
    sheet = workbook["期初库存导入"]
    for row in rows:
        sheet.append(row)
    output = io.BytesIO()
    workbook.save(output)
    workbook.close()
    return SimpleUploadedFile(
        "opening-stock.xlsx",
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


def test_opening_stock_template_has_fixed_headers_validations_and_numeric_example():
    company = make_company()
    data = build_template_workbook("opening_stock", company=company)
    workbook = load_workbook(io.BytesIO(data))
    sheet = workbook["期初库存导入"]
    assert [cell.value for cell in sheet[1]] == [
        "公司编码",
        "仓库编码",
        "物品编码",
        "数量",
        "单位成本",
        "0成本原因",
        "备注",
    ]
    assert isinstance(workbook["示例"]["D2"].value, int)
    assert isinstance(workbook["示例"]["E2"].value, int)
    assert len(sheet.data_validations.dataValidation) == 2
    workbook.close()


def test_opening_import_reports_company_master_decimal_and_zero_cost_errors():
    company = make_company()
    actor = make_user("s14-opening-invalid", "warehouse")
    category = make_supply_category(company)
    warehouse = make_supply_warehouse(company, "WH-A")
    item = make_supply_item(company, category, "ITEM-A")
    batch = upload_and_validate_import(
        actor=actor,
        company=company,
        import_type="opening_stock",
        uploaded_file=opening_workbook(
            company,
            [
                ["OTHER", warehouse.code, item.item_code, 1, 1, "", ""],
                [company.code, "MISSING", item.item_code, 1, 1, "", ""],
                [company.code, warehouse.code, "MISSING", 1, 1, "", ""],
                [company.code, warehouse.code, item.item_code, 0, 1, "", ""],
                [company.code, warehouse.code, item.item_code, 1, 0, "", ""],
            ],
        ),
        idempotency_key="s14-opening-invalid",
    )
    assert batch.status == "invalid"
    assert batch.error_rows == 5
    assert not SupplyDocument.objects.exists()
    errors = [error for row in batch.rows.all() for error in row.errors_json]
    assert {error["field"] for error in errors} >= {
        "公司编码",
        "仓库编码",
        "物品编码",
        "数量",
        "0成本原因",
    }


def test_opening_import_confirmation_groups_by_warehouse_creates_only_drafts_and_is_idempotent():
    company = make_company()
    actor = make_user("s14-opening-confirm", "finance")
    category = make_supply_category(company)
    warehouse_a = make_supply_warehouse(company, "WH-A")
    warehouse_b = make_supply_warehouse(company, "WH-B")
    item_a = make_supply_item(company, category, "ITEM-A")
    item_b = make_supply_item(company, category, "ITEM-B")
    batch = upload_and_validate_import(
        actor=actor,
        company=company,
        import_type="opening_stock",
        uploaded_file=opening_workbook(
            company,
            [
                [company.code, warehouse_a.code, item_a.item_code, 10, 100, "", "A"],
                [company.code, warehouse_a.code, item_b.item_code, 2, 0, "赠品", "B"],
                [company.code, warehouse_b.code, item_a.item_code, 3, 120, "", "C"],
            ],
        ),
        idempotency_key="s14-opening-confirm",
    )
    assert batch.status == "validated"
    confirm_import_batch(actor=actor, batch=batch)
    batch.refresh_from_db()
    assert batch.status == "confirmed"
    assert SupplyDocument.objects.filter(
        company=company, document_type="opening", status="draft"
    ).count() == 2
    assert SupplyDocument.objects.filter(target_warehouse=warehouse_a).get().lines.count() == 2
    assert SupplyDocument.objects.filter(target_warehouse=warehouse_b).get().lines.count() == 1
    assert batch.rows.filter(
        validation_status="created", created_object_type="SupplyDocumentLine"
    ).count() == 3
    assert not SupplyStockBalance.objects.exists()
    assert not SupplyStockLedger.objects.exists()

    confirm_import_batch(actor=actor, batch=batch)
    assert SupplyDocument.objects.count() == 2


def test_opening_import_confirmation_change_rolls_back_all_documents():
    company = make_company()
    actor = make_user("s14-opening-race", "system_admin")
    category = make_supply_category(company)
    warehouse = make_supply_warehouse(company)
    item = make_supply_item(company, category)
    batch = upload_and_validate_import(
        actor=actor,
        company=company,
        import_type="opening_stock",
        uploaded_file=opening_workbook(
            company,
            [[company.code, warehouse.code, item.item_code, 1, 10, "", ""]],
        ),
        idempotency_key="s14-opening-race",
    )
    item.is_active = False
    item.save(update_fields=["is_active"])
    with pytest.raises(ValidationError):
        confirm_import_batch(actor=actor, batch=batch)
    batch.refresh_from_db()
    assert batch.status == "validated"
    assert not SupplyDocument.objects.exists()
    assert not batch.rows.filter(validation_status="created").exists()


def test_opening_import_http_permissions_match_document_action(client):
    make_company()
    warehouse = make_user("s14-opening-http-warehouse", "warehouse")
    client.force_login(warehouse)
    assert client.get(reverse("imports:template", args=["opening_stock"])).status_code == 200
    assert client.get(reverse("imports:upload", args=["opening_stock"])).status_code == 200

    for role in ("equipment", "management", "employee"):
        actor = make_user(f"s14-opening-http-{role}", role)
        client.force_login(actor)
        assert client.get(reverse("imports:template", args=["opening_stock"])).status_code == 403
        assert client.get(reverse("imports:upload", args=["opening_stock"])).status_code == 403
