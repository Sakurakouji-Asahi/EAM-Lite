import io
import re
import zipfile
from unittest.mock import patch

import pytest
from django.contrib.auth import get_user_model
from django.contrib.auth.models import Group
from django.core.exceptions import PermissionDenied, ValidationError
from django.core.files.uploadedfile import SimpleUploadedFile
from openpyxl import load_workbook

from apps.audit.models import AuditLog
from apps.imports import services as import_services
from apps.imports.services import (
    build_template_workbook,
    confirm_import_batch,
    upload_and_validate_import,
)
from apps.masterdata.models import Company, Department, Employee, ImportBatch


pytestmark = pytest.mark.django_db


@pytest.fixture(autouse=True)
def isolated_import_media(settings, tmp_path):
    """Keep uploaded source workbooks out of the repository media directory."""
    settings.MEDIA_ROOT = tmp_path / "media"


def setup_data(role):
    owner = Company.objects.create(
        code="C1", normalized_code="c1", name="测试公司", short_name="测试"
    )
    user = get_user_model().objects.create_user(
        username=role,
        password="Valid-Password-2026!",
        display_name=role,
    )
    user.groups.add(Group.objects.get(name=role))
    return owner, user


def workbook_file(import_type, rows, *, formula=False, unknown_header=False):
    data = build_template_workbook(import_type)
    book = load_workbook(io.BytesIO(data))
    sheet = book["部门导入" if import_type == "department" else "人员导入"]
    if unknown_header:
        sheet.cell(1, 1).value = "未知列"
    for row in rows:
        sheet.append(row)
    if formula:
        sheet.cell(2, 2).value = "=1+1"
    output = io.BytesIO()
    book.save(output)
    return SimpleUploadedFile(
        f"{import_type}.xlsx",
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


def workbook_bytes(import_type, mutate):
    data = build_template_workbook(import_type)
    book = load_workbook(io.BytesIO(data))
    mutate(book)
    output = io.BytesIO()
    book.save(output)
    book.close()
    return output.getvalue()


def uploaded_workbook(data):
    return SimpleUploadedFile(
        "department.xlsx",
        data,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


def rewrite_zip_member(data, member_name, transform):
    source = zipfile.ZipFile(io.BytesIO(data))
    output = io.BytesIO()
    with source, zipfile.ZipFile(output, "w", zipfile.ZIP_DEFLATED) as target:
        for item in source.infolist():
            content = source.read(item)
            if item.filename == member_name:
                content = transform(content)
            target.writestr(item, content)
    return output.getvalue()


def test_department_import_preview_errors_do_not_write_business_rows():
    owner, admin = setup_data("system_admin")
    upload = workbook_file(
        "department",
        [
            ["D1", "部门一", "MISSING", "", "是"],
            ["D1", "重复部门", "", "", "是"],
        ],
    )

    batch = upload_and_validate_import(
        actor=admin,
        company=owner,
        import_type="department",
        uploaded_file=upload,
        idempotency_key="11111111-1111-1111-1111-111111111111",
    )

    assert batch.status == "invalid"
    assert batch.error_rows == 2
    assert not Department.objects.filter(company=owner).exists()
    errors = [error for row in batch.rows.all() for error in row.errors_json]
    assert all({"field", "value", "reason"} <= set(error) for error in errors)


def test_mixed_valid_and_invalid_preview_writes_no_masterdata():
    owner, admin = setup_data("system_admin")
    upload = workbook_file(
        "department",
        [
            ["GOOD", "正确部门", "", "", "是"],
            ["BAD", "错误部门", "MISSING", "", "是"],
        ],
    )

    batch = upload_and_validate_import(
        actor=admin,
        company=owner,
        import_type="department",
        uploaded_file=upload,
        idempotency_key="mixed-valid-invalid-preview",
    )

    assert batch.status == "invalid"
    assert batch.total_rows == 2
    assert batch.valid_rows == 1
    assert batch.error_rows == 1
    assert list(
        batch.rows.order_by("row_number").values_list(
            "validation_status", flat=True
        )
    ) == ["valid", "invalid"]
    assert not Department.objects.filter(company=owner).exists()


def test_department_import_confirm_is_atomic_and_audited():
    owner, admin = setup_data("system_admin")
    upload = workbook_file(
        "department",
        [
            ["ROOT", "根部门", "", "", "是"],
            ["CHILD", "子部门", "ROOT", "", "是"],
        ],
    )
    batch = upload_and_validate_import(
        actor=admin,
        company=owner,
        import_type="department",
        uploaded_file=upload,
        idempotency_key="22222222-2222-2222-2222-222222222222",
    )
    assert batch.status == "validated"

    confirm_import_batch(actor=admin, batch=batch)
    batch.refresh_from_db()
    assert batch.status == "confirmed"
    assert batch.rows.filter(validation_status="created").count() == 2
    assert Department.objects.get(normalized_code="child").parent.normalized_code == "root"
    assert AuditLog.objects.filter(
        company=owner, action="import_confirm", object_id=str(batch.pk)
    ).exists()
    audit = AuditLog.objects.get(
        company=owner, action="import_confirm", object_id=str(batch.pk)
    )
    assert audit.old_data_json == {"status": "validated"}
    assert audit.new_data_json["status"] == "confirmed"
    assert len(audit.new_data_json["created_objects"]) == 2
    assert confirm_import_batch(actor=admin, batch=batch).pk == batch.pk
    assert Department.objects.filter(company=owner).count() == 2


def test_employee_import_unknown_department_invalid_and_unauthorized_rejected():
    owner, hr = setup_data("hr")
    ordinary = get_user_model().objects.create_user(
        username="ordinary",
        password="Valid-Password-2026!",
        display_name="普通用户",
    )
    upload = workbook_file(
        "employee",
        [["E1", "张三", "UNKNOWN", "active", "2026-01-01", "", "", "", "是"]],
    )
    batch = upload_and_validate_import(
        actor=hr,
        company=owner,
        import_type="employee",
        uploaded_file=upload,
        idempotency_key="33333333-3333-3333-3333-333333333333",
    )
    assert batch.status == "invalid"
    assert not Employee.objects.exists()

    with pytest.raises(PermissionDenied):
        upload_and_validate_import(
            actor=ordinary,
            company=owner,
            import_type="employee",
            uploaded_file=workbook_file("employee", []),
            idempotency_key="44444444-4444-4444-4444-444444444444",
        )


def test_employee_import_confirm_success_is_audited_and_idempotent():
    owner, hr = setup_data("hr")
    Department.objects.create(
        company=owner,
        code="D1",
        normalized_code="d1",
        name="部门一",
    )
    upload = workbook_file(
        "employee",
        [
            ["E1", "张三", "D1", "active", "2026-01-01", "", "", "", "是"],
            [
                "E2",
                "李四",
                "D1",
                "resigned",
                "2025-01-01",
                "2026-01-31",
                "",
                "",
                "否",
            ],
        ],
    )
    batch = upload_and_validate_import(
        actor=hr,
        company=owner,
        import_type="employee",
        uploaded_file=upload,
        idempotency_key="88888888-8888-8888-8888-888888888888",
    )
    assert batch.status == "validated"

    confirm_import_batch(actor=hr, batch=batch)
    assert Employee.objects.filter(company=owner).count() == 2
    resigned = Employee.objects.get(normalized_employee_no="e2")
    assert resigned.employment_status == "resigned"
    assert not resigned.is_active
    assert AuditLog.objects.filter(
        company=owner, action="import_confirm", object_id=str(batch.pk)
    ).exists()
    confirm_import_batch(actor=hr, batch=batch)
    assert Employee.objects.filter(company=owner).count() == 2


def test_employee_confirm_failure_rolls_back_all_objects_and_audit():
    owner, hr = setup_data("hr")
    Department.objects.create(
        company=owner,
        code="D1",
        normalized_code="d1",
        name="部门一",
    )
    upload = workbook_file(
        "employee",
        [
            ["E1", "张三", "D1", "active", "", "", "", "", "是"],
            ["E2", "李四", "D1", "active", "", "", "", "", "是"],
        ],
    )
    batch = upload_and_validate_import(
        actor=hr,
        company=owner,
        import_type="employee",
        uploaded_file=upload,
        idempotency_key="99999999-9999-9999-9999-999999999999",
    )
    real_create_employee = import_services.create_employee
    calls = 0

    def fail_on_second_row(**kwargs):
        nonlocal calls
        calls += 1
        if calls == 2:
            raise RuntimeError("simulated employee write failure")
        return real_create_employee(**kwargs)

    with patch(
        "apps.imports.services.create_employee", side_effect=fail_on_second_row
    ):
        with pytest.raises(RuntimeError):
            confirm_import_batch(actor=hr, batch=batch)

    batch.refresh_from_db()
    assert batch.status == "validated"
    assert not Employee.objects.filter(company=owner).exists()
    assert not batch.rows.filter(validation_status="created").exists()
    assert not AuditLog.objects.filter(
        action="import_confirm", object_id=str(batch.pk)
    ).exists()


def test_import_rejects_invalid_date_and_database_duplicate():
    owner, hr = setup_data("hr")
    department = Department.objects.create(
        company=owner,
        code="D1",
        normalized_code="d1",
        name="部门一",
    )
    Employee.objects.create(
        company=owner,
        department=department,
        employee_no="EXISTING",
        normalized_employee_no="existing",
        name="现有员工",
    )
    batch = upload_and_validate_import(
        actor=hr,
        company=owner,
        import_type="employee",
        uploaded_file=workbook_file(
            "employee",
            [
                [
                    "EXISTING",
                    "重复员工",
                    "D1",
                    "active",
                    "2026-02-30",
                    "",
                    "",
                    "",
                    "是",
                ]
            ],
        ),
        idempotency_key="aaaaaaaa-aaaa-aaaa-aaaa-aaaaaaaaaaaa",
    )
    errors = batch.rows.get().errors_json
    assert batch.status == "invalid"
    assert any(error["field"] == "员工编号" for error in errors)
    assert any(error["field"] == "入职日期" for error in errors)


def test_confirm_view_requires_explicit_checkbox(client):
    owner, admin = setup_data("system_admin")
    batch = upload_and_validate_import(
        actor=admin,
        company=owner,
        import_type="department",
        uploaded_file=workbook_file(
            "department", [["D1", "部门一", "", "", "是"]]
        ),
        idempotency_key="bbbbbbbb-bbbb-bbbb-bbbb-bbbbbbbbbbbb",
    )
    client.force_login(admin)

    response = client.post(f"/imports/batches/{batch.pk}/confirm/", {})
    batch.refresh_from_db()
    assert response.status_code == 302
    assert batch.status == "validated"
    assert not Department.objects.filter(company=owner).exists()

    response = client.post(
        f"/imports/batches/{batch.pk}/confirm/", {"confirm": "1"}
    )
    batch.refresh_from_db()
    assert response.status_code == 302
    assert batch.status == "confirmed"
    assert Department.objects.filter(company=owner).count() == 1


@pytest.mark.parametrize("options", [{"formula": True}, {"unknown_header": True}])
def test_import_rejects_formula_and_unknown_columns_as_structured_errors(options):
    owner, admin = setup_data("system_admin")
    upload = workbook_file(
        "department", [["D1", "部门", "", "", "是"]], **options
    )
    batch = upload_and_validate_import(
        actor=admin,
        company=owner,
        import_type="department",
        uploaded_file=upload,
        idempotency_key=(
            "55555555-5555-5555-5555-555555555555"
            if options.get("formula")
            else "66666666-6666-6666-6666-666666666666"
        ),
    )
    assert batch.status == "invalid"
    assert batch.rows.filter(validation_status="invalid").exists()


def test_confirm_failure_rolls_back_objects_rows_batch_and_audit():
    owner, admin = setup_data("system_admin")
    upload = workbook_file(
        "department", [["D1", "部门一", "", "", "是"], ["D2", "部门二", "", "", "是"]]
    )
    batch = upload_and_validate_import(
        actor=admin,
        company=owner,
        import_type="department",
        uploaded_file=upload,
        idempotency_key="77777777-7777-7777-7777-777777777777",
    )

    real_create_department = import_services.create_department
    calls = 0

    def fail_on_second_row(**kwargs):
        nonlocal calls
        calls += 1
        if calls == 2:
            raise RuntimeError("simulated write failure")
        return real_create_department(**kwargs)

    with patch(
        "apps.imports.services.create_department",
        side_effect=fail_on_second_row,
    ):
        with pytest.raises(RuntimeError):
            confirm_import_batch(actor=admin, batch=batch)

    batch.refresh_from_db()
    assert batch.status == "validated"
    assert not batch.rows.filter(validation_status="created").exists()
    assert not Department.objects.filter(company=owner).exists()
    assert not AuditLog.objects.filter(action="import_confirm", object_id=str(batch.pk)).exists()


def test_import_rejects_sparse_dimension_before_openpyxl_expansion():
    owner, admin = setup_data("system_admin")
    data = workbook_bytes(
        "department",
        lambda book: setattr(book["部门导入"]["A1000000"], "value", "SPARSE"),
    )

    batch = upload_and_validate_import(
        actor=admin,
        company=owner,
        import_type="department",
        uploaded_file=uploaded_workbook(data),
        idempotency_key="sparse-dimension-rejected",
    )

    assert batch.status == "invalid"
    assert "实际单元格超过安全上限" in str(batch.rows.get().errors_json)


def test_import_rejects_formula_in_instructions_sheet():
    owner, admin = setup_data("system_admin")
    data = workbook_bytes(
        "department",
        lambda book: setattr(book["填写说明"]["B3"], "value", "=1+1"),
    )

    batch = upload_and_validate_import(
        actor=admin,
        company=owner,
        import_type="department",
        uploaded_file=uploaded_workbook(data),
        idempotency_key="instruction-formula-rejected",
    )

    assert batch.status == "invalid"
    assert "所有工作表均禁止公式" in str(batch.rows.get().errors_json)


def test_import_rejects_formula_hidden_beyond_forged_dimension():
    owner, admin = setup_data("system_admin")
    data = workbook_bytes(
        "department",
        lambda book: setattr(book["填写说明"]["B3"], "value", "=1+1"),
    )
    data = rewrite_zip_member(
        data,
        "xl/worksheets/sheet2.xml",
        lambda content: re.sub(
            rb'<dimension ref="[^"]+"\s*/>',
            b'<dimension ref="A1:B2"/>',
            content,
            count=1,
        ),
    )

    batch = upload_and_validate_import(
        actor=admin,
        company=owner,
        import_type="department",
        uploaded_file=uploaded_workbook(data),
        idempotency_key="formula-forged-dimension-rejected",
    )

    assert batch.status == "invalid"
    assert "所有工作表均禁止公式" in str(batch.rows.get().errors_json)


def test_import_rejects_far_cell_hidden_beyond_forged_dimension():
    owner, admin = setup_data("system_admin")
    data = workbook_bytes(
        "department",
        lambda book: setattr(book["部门导入"]["A1000000"], "value", "SPARSE"),
    )
    data = rewrite_zip_member(
        data,
        "xl/worksheets/sheet1.xml",
        lambda content: re.sub(
            rb'<dimension ref="[^"]+"\s*/>',
            b'<dimension ref="A1:E2"/>',
            content,
            count=1,
        ),
    )

    batch = upload_and_validate_import(
        actor=admin,
        company=owner,
        import_type="department",
        uploaded_file=uploaded_workbook(data),
        idempotency_key="far-cell-forged-dimension-rejected",
    )

    assert batch.status == "invalid"
    assert "实际单元格超过安全上限" in str(batch.rows.get().errors_json)


def test_import_rejects_external_relationship_whitespace_variant():
    owner, admin = setup_data("system_admin")
    base = build_template_workbook("department")
    source = zipfile.ZipFile(io.BytesIO(base))
    output = io.BytesIO()
    with source, zipfile.ZipFile(output, "w", zipfile.ZIP_DEFLATED) as target:
        for item in source.infolist():
            content = source.read(item)
            if item.filename == "xl/_rels/workbook.xml.rels":
                marker = b"</Relationships>"
                relationship = (
                    b'<Relationship Id="rIdExternal" '
                    b'Type="http://schemas.openxmlformats.org/officeDocument/2006/'
                    b'relationships/hyperlink" Target="https://example.invalid/" '
                    b'tArGeTmOdE \n = \n " ExTeRnAl "/>'
                )
                content = content.replace(marker, relationship + marker)
            target.writestr(item, content)

    with pytest.raises(ValidationError, match="外部关系"):
        upload_and_validate_import(
            actor=admin,
            company=owner,
            import_type="department",
            uploaded_file=uploaded_workbook(output.getvalue()),
            idempotency_key="external-relation-whitespace-rejected",
        )


def test_xlsx_preflight_limits_actual_cell_count(monkeypatch):
    monkeypatch.setattr(import_services, "MAX_WORKSHEET_CELLS", 2)
    worksheet = (
        b'<worksheet xmlns="http://schemas.openxmlformats.org/'
        b'spreadsheetml/2006/main"><sheetData><row r="1">'
        b'<c r="A1"/><c r="B1"/><c r="C1"/>'
        b"</row></sheetData></worksheet>"
    )

    errors = import_services._validate_worksheet_xml(
        io.BytesIO(worksheet), "xl/worksheets/sheet1.xml"
    )

    assert "实际单元格数量超过安全上限" in str(errors)


def test_xlsx_preflight_rejects_oversized_member_without_large_fixture(monkeypatch):
    monkeypatch.setattr(import_services, "MAX_ARCHIVE_MEMBER_BYTES", 256)
    base = build_template_workbook("department")
    source = zipfile.ZipFile(io.BytesIO(base))
    output = io.BytesIO()
    with source, zipfile.ZipFile(output, "w", zipfile.ZIP_DEFLATED) as target:
        for item in source.infolist():
            target.writestr(item, source.read(item))
        target.writestr("docProps/oversized.bin", b"A" * 257)

    with pytest.raises(ValidationError, match="单个内部文件解压后过大"):
        import_services._validate_xlsx_container(output.getvalue())


def test_xlsx_preflight_rejects_high_compression_ratio():
    base = build_template_workbook("department")
    source = zipfile.ZipFile(io.BytesIO(base))
    output = io.BytesIO()
    with source, zipfile.ZipFile(output, "w", zipfile.ZIP_DEFLATED) as target:
        for item in source.infolist():
            target.writestr(item, source.read(item))
        target.writestr("docProps/compression-bomb.bin", b"0" * 100_000)

    with pytest.raises(ValidationError, match="压缩比异常"):
        import_services._validate_xlsx_container(output.getvalue())
