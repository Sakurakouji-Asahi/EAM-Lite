from __future__ import annotations

import io
import json
from decimal import Decimal

import pytest
from django.core.exceptions import PermissionDenied, ValidationError
from openpyxl import load_workbook

from apps.assets.models import Asset
from apps.audit.models import AuditLog
from apps.assets.services import create_asset_draft
from apps.imports.services import (
    build_template_workbook,
    get_template_definition,
    upload_and_validate_import,
)
from apps.masterdata.models import UserDepartmentScope
from tests.test_sprint3_support import (
    make_category,
    make_company,
    make_department,
    make_employee,
    make_location_tree,
    make_user,
)
from tests.test_sprint5_support import (
    XLSX_MIME,
    add_finance_row,
    asset_workbook_upload,
    finance_configuration,
    physical_row,
    sprint5_context,
)


pytestmark = pytest.mark.django_db


@pytest.fixture(autouse=True)
def isolated_import_media(settings, tmp_path):
    settings.MEDIA_ROOT = tmp_path / "media"


def _upload(actor, company, workbook, key):
    return upload_and_validate_import(
        actor=actor,
        company=company,
        import_type="asset_initialization",
        uploaded_file=workbook,
        idempotency_key=key,
    )


def _assert_invalid_batch_audited(batch):
    assert AuditLog.objects.filter(
        company=batch.company,
        action="import_failure",
        object_type="ImportBatch",
        object_id=str(batch.pk),
    ).exists()


def _all_messages(batch):
    return [
        message
        for row in batch.rows.order_by("row_number")
        for message in [*row.errors_json, *row.warnings_json]
    ]


def test_versioned_chinese_template_has_instructions_example_and_exact_match_keys():
    company, _actor, *_ = sprint5_context(prefix="S5T")
    definition = get_template_definition("asset_initialization", company=company)
    content = build_template_workbook("asset_initialization", company=company)

    workbook = load_workbook(io.BytesIO(content), data_only=False)
    assert workbook.sheetnames == [definition.sheet_name, "填写说明", "示例"]
    assert workbook["填写说明"]["A2"].value == "模板版本"
    assert workbook["填写说明"]["B2"].value == definition.version
    assert tuple(cell.value for cell in workbook[definition.sheet_name][1]) == definition.headers
    instructions = "\n".join(
        str(cell.value or "") for row in workbook["填写说明"] for cell in row
    )
    assert "employee_no" in instructions
    assert "code" in instructions
    assert "仅创建草稿" in instructions
    assert "本机路径" in instructions
    workbook.close()


@pytest.mark.parametrize(
    ("filename", "mime"),
    (("assets.xls", XLSX_MIME), ("assets.xlsx", "text/plain")),
)
def test_import_rejects_wrong_extension_or_mime(filename, mime):
    company, actor, category, department, employee, location = sprint5_context(
        prefix=f"S5SEC{len(filename)}{len(mime)}"
    )
    upload = asset_workbook_upload(
        company,
        [physical_row(company, category, department, employee, location)],
        filename=filename,
        content_type=mime,
    )
    with pytest.raises(ValidationError):
        _upload(actor, company, upload, f"bad-container-{filename}-{mime}")
    assert not Asset.objects.filter(company=company).exists()


@pytest.mark.parametrize("mutation", ("sheet", "version", "missing", "unknown", "duplicate", "formula"))
def test_workbook_contract_errors_are_structured_and_never_write_assets(mutation):
    company, actor, category, department, employee, location = sprint5_context(
        prefix=f"S5C{mutation[:3].upper()}"
    )

    def mutate(workbook, sheet, _definition):
        if mutation == "sheet":
            sheet.title = "错误工作表"
        elif mutation == "version":
            workbook["填写说明"]["B2"] = "obsolete-v0"
        elif mutation == "missing":
            sheet.delete_cols(1)
        elif mutation == "unknown":
            sheet.cell(1, 1).value = "未知列"
        elif mutation == "duplicate":
            sheet.cell(1, 2).value = sheet.cell(1, 1).value
        else:
            sheet.cell(2, 1).value = "=1+1"

    upload = asset_workbook_upload(
        company,
        [physical_row(company, category, department, employee, location)],
        mutate=mutate,
    )
    batch = _upload(actor, company, upload, f"contract-{mutation}")

    assert batch.status == "invalid"
    assert batch.error_rows >= 1
    _assert_invalid_batch_audited(batch)
    assert not Asset.objects.filter(company=company).exists()
    assert all(
        {"field", "value", "reason"} <= set(error)
        for row in batch.rows.all()
        for error in row.errors_json
    )


def test_invalid_business_row_pinpoints_row_field_value_reason_without_writes():
    company, actor, category, department, employee, location = sprint5_context(
        prefix="S5ERR"
    )
    row = physical_row(
        company,
        category,
        department,
        employee,
        location,
        **{
            "数量": 2,
            "购置日期": "2026-02-30",
            "是否需要保养": "maybe",
            "责任员工编号": "MISSING",
        },
    )
    batch = _upload(
        actor,
        company,
        asset_workbook_upload(company, [row]),
        "invalid-row-details",
    )

    assert batch.status == "invalid"
    staged = batch.rows.get()
    assert staged.row_number == 2
    assert {item["field"] for item in staged.errors_json} >= {
        "数量",
        "购置日期",
        "是否需要保养",
        "责任员工编号",
    }
    assert all({"field", "value", "reason"} == set(item) for item in staged.errors_json)
    assert not Asset.objects.filter(company=company).exists()


def test_cross_company_references_and_department_manager_outside_scope_are_rejected():
    company, manager, category, inside, employee, location = sprint5_context(
        role="department_manager", prefix="S5SCOPE"
    )
    UserDepartmentScope.objects.create(
        company=company,
        user=manager,
        department=inside,
        include_descendants=True,
        assigned_by=manager,
    )
    outside = make_department(company, "S5-OUT")
    outside_employee = make_employee(company, outside, "S5-OUT-E")
    row = physical_row(
        company,
        category,
        outside,
        outside_employee,
        location,
    )
    batch = _upload(
        manager,
        company,
        asset_workbook_upload(company, [row]),
        "department-outside-scope",
    )
    assert batch.status == "invalid"
    assert "范围" in json.dumps(_all_messages(batch), ensure_ascii=False)

    other = make_company("S5OTHER", active=False)
    other_category = make_category(other, "OTHER-EQ")
    other_department = make_department(other, "OTHER-D")
    other_employee = make_employee(other, other_department, "OTHER-E")
    _site, _area, other_location = make_location_tree(other, "OTHER-L")
    cross_row = physical_row(
        company,
        other_category,
        other_department,
        other_employee,
        other_location,
    )
    cross_batch = _upload(
        manager,
        company,
        asset_workbook_upload(company, [cross_row]),
        "cross-company-references",
    )
    assert cross_batch.status == "invalid"
    assert not Asset.objects.exists()


def test_file_and_database_potential_duplicates_are_warnings_not_business_writes():
    company, actor, category, department, employee, location = sprint5_context(
        prefix="S5DUP"
    )
    create_asset_draft(
        actor=actor,
        company=company,
        data={
            "asset_name": "现有设备",
            "category": category,
            "department": department,
            "responsible_employee": employee,
            "location": location,
            "quantity": 1,
            "unit": "台",
            "serial_number": "DUP-SERIAL",
            "factory_number": "DUP-FACTORY",
            "historical_code": "DUP-HISTORY",
        },
    )
    first = physical_row(
        company,
        category,
        department,
        employee,
        location,
        **{
            "资产名称": "导入一",
            "序列号": "DUP-SERIAL",
            "出厂编号": "DUP-FACTORY",
            "历史参考编号": "DUP-HISTORY",
        },
    )
    second = physical_row(
        company,
        category,
        department,
        employee,
        location,
        **{
            "资产名称": "导入二",
            "序列号": "DUP-SERIAL",
            "出厂编号": "OTHER-FACTORY",
            "历史参考编号": "OTHER-HISTORY",
        },
    )
    batch = _upload(
        actor,
        company,
        asset_workbook_upload(company, [first, second]),
        "duplicate-warnings",
    )

    assert batch.status == "validated"
    assert batch.warning_rows == 2
    assert batch.rows.exclude(warnings_json=[]).count() == 2
    assert Asset.objects.filter(company=company).count() == 1


def test_non_finance_financial_columns_are_explicit_errors_and_threshold_is_only_warning():
    company, equipment, category, department, employee, location = sprint5_context(
        prefix="S5FINBOUND"
    )
    forbidden = physical_row(company, category, department, employee, location)
    forbidden["原值"] = "5000.00"
    denied = _upload(
        equipment,
        company,
        asset_workbook_upload(company, [forbidden]),
        "equipment-financial-column",
    )
    assert denied.status == "invalid"
    assert "财务" in json.dumps(_all_messages(denied), ensure_ascii=False)
    assert not Asset.objects.filter(company=company).exists()

    finance = make_user("s5-threshold-finance", "finance")
    controlled = physical_row(company, category, department, employee, location)
    controlled.update(
        {
            "会计认定": "controlled_non_fixed",
            "会计认定说明": "达到提示金额但经财务判断不资本化",
            "原值": "5000.00",
            "资本化日期": "2024-01-01",
        }
    )
    warning_batch = _upload(
        finance,
        company,
        asset_workbook_upload(company, [controlled]),
        "threshold-warning-only",
    )
    assert warning_batch.status == "validated"
    assert warning_batch.error_rows == 0
    assert warning_batch.warning_rows == 1
    normalized = warning_batch.rows.get().normalized_data_json
    assert normalized["finance_data"]["accounting_treatment"] == "controlled_non_fixed"
    assert Decimal(normalized["finance_data"]["original_cost"]) == Decimal("5000.00")
    assert not Asset.objects.filter(company=company).exists()


def test_opening_balances_are_decimal_and_theoretical_reference_does_not_replace_actual():
    company, finance, category, department, employee, location = sprint5_context(
        role="finance", prefix="S5OPEN"
    )
    fixed, policy = finance_configuration(company, finance)
    row = physical_row(company, category, department, employee, location)
    add_finance_row(
        row,
        fixed_category=fixed,
        policy=policy,
        opening_ad="1234.56",
        opening_book="10765.44",
    )
    batch = _upload(
        finance,
        company,
        asset_workbook_upload(company, [row]),
        "opening-balance-preview",
    )
    assert batch.status == "validated"
    normalized = batch.rows.get().normalized_data_json
    profile = normalized["profile_data"]
    theoretical = normalized["theoretical_reference"]
    assert Decimal(profile["opening_actual_accumulated_depreciation"]) == Decimal(
        "1234.56"
    )
    assert Decimal(profile["opening_book_value"]) == Decimal("10765.44")
    # The independent run starts from original cost rather than the imported
    # actual opening book value; its result therefore exceeds the same run
    # incorrectly seeded from 10,765.44.
    # as_of_date includes the complete December period ending on 2026-01-01.
    assert Decimal(theoretical["theoretical_book_value"]) == Decimal("7440.00")
    assert Decimal(theoretical["planned_accumulated_depreciation"]) != Decimal(
        profile["opening_actual_accumulated_depreciation"]
    )
    assert not Asset.objects.filter(company=company).exists()


def test_native_excel_numeric_money_uses_raw_decimal_literal_without_float_rejection():
    company, finance, category, department, employee, location = sprint5_context(
        role="finance", prefix="S5NUMERIC"
    )
    fixed, policy = finance_configuration(company, finance)
    row = physical_row(company, category, department, employee, location)
    add_finance_row(
        row,
        fixed_category=fixed,
        policy=policy,
        cost=1234.56,
        opening_ad=234.56,
        opening_book=1000.00,
    )
    batch = _upload(
        finance,
        company,
        asset_workbook_upload(company, [row]),
        "native-excel-numeric-money",
    )
    assert batch.status == "validated", batch.rows.get().errors_json
    normalized = batch.rows.get().normalized_data_json
    assert Decimal(normalized["finance_data"]["original_cost"]) == Decimal("1234.56")
    assert Decimal(
        normalized["profile_data"]["opening_actual_accumulated_depreciation"]
    ) == Decimal("234.56")


@pytest.mark.parametrize(
    ("expected_units", "work_unit", "valid"),
    (("", "", False), ("10000.0000", "台时", True)),
)
def test_units_of_production_requires_total_units_and_work_unit(
    expected_units, work_unit, valid
):
    company, finance, category, department, employee, location = sprint5_context(
        role="finance", prefix=f"S5UOP{'OK' if valid else 'BAD'}"
    )
    fixed, policy = finance_configuration(
        company, finance, method="units_of_production", key=f"S5-UOP-{valid}"
    )
    row = physical_row(company, category, department, employee, location)
    add_finance_row(
        row,
        fixed_category=fixed,
        policy=policy,
        method="units_of_production",
        **{"预计总工作量": expected_units, "工作量单位": work_unit},
    )
    batch = _upload(
        finance,
        company,
        asset_workbook_upload(company, [row]),
        f"uop-fields-{valid}",
    )
    assert (batch.status == "validated") is valid
    assert (batch.error_rows == 0) is valid
