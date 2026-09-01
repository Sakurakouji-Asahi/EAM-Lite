from datetime import date
from decimal import Decimal
from io import BytesIO

import pytest
from django.core.exceptions import PermissionDenied
from django.db import connection
from django.test.utils import CaptureQueriesContext
from django.urls import reverse
from django.utils import timezone
from openpyxl import load_workbook

from apps.reports.excel import write_report_workbook
from apps.reports.models import ExportLog
from apps.reports.queries import build_report_dataset
from apps.reports.services import generate_report_export
from apps.reports.permissions import can_download_export
from apps.masterdata.models import UserDepartmentScope
from apps.supplies.models import SupplyCustody
from apps.supplies.services import (
    create_supply_count_task,
    post_supply_document,
    publish_supply_count_task,
    record_supply_count,
    reverse_supply_document,
)
from tests.test_sprint6_support import formal_asset_context
from tests.test_sprint15_support import (
    make_company,
    make_department,
    make_employee,
    make_issue_document,
    make_supply_category,
    make_supply_item,
    make_supply_warehouse,
    make_user,
    seed_supply_stock,
)


pytestmark = pytest.mark.django_db


def report_context():
    company = make_company("S18R")
    warehouse_user = make_user("s18-report-warehouse", "warehouse")
    employee_user = make_user("s18-report-employee", "employee")
    department = make_department(company, "S18USE")
    employee = make_employee(
        company, department, "S18EMP", user=employee_user
    )
    category = make_supply_category(company, "S18CAT")
    warehouse = make_supply_warehouse(company, "S18WH")
    paper = make_supply_item(
        company,
        category,
        "S18PAPER",
        unit="箱",
        minimum_stock_quantity=Decimal("10.0000"),
        default_warehouse=warehouse,
    )
    chair = make_supply_item(
        company,
        category,
        "S18CHAIR",
        item_type="durable_quantity",
        unit="把",
    )
    seed_supply_stock(
        actor=warehouse_user,
        company=company,
        warehouse=warehouse,
        item=paper,
        quantity="12",
        unit_cost="100",
        key="s18-paper-opening",
    )
    paper_issue = make_issue_document(
        actor=warehouse_user,
        company=company,
        warehouse=warehouse,
        item=paper,
        department=department,
        employee=employee,
        quantity="5",
        key="s18-paper-issue",
    )
    post_supply_document(document=paper_issue, actor=warehouse_user)
    seed_supply_stock(
        actor=warehouse_user,
        company=company,
        warehouse=warehouse,
        item=chair,
        quantity="3",
        unit_cost="80",
        key="s18-chair-opening",
    )
    chair_issue = make_issue_document(
        actor=warehouse_user,
        company=company,
        warehouse=warehouse,
        item=chair,
        department=department,
        employee=employee,
        quantity="1",
        key="s18-chair-issue",
    )
    post_supply_document(document=chair_issue, actor=warehouse_user)
    return {
        "company": company,
        "warehouse_user": warehouse_user,
        "employee_user": employee_user,
        "department": department,
        "employee": employee,
        "category": category,
        "warehouse": warehouse,
        "paper": paper,
        "chair": chair,
        "paper_issue": paper_issue,
        "chair_issue": chair_issue,
    }


def rows(context, report_key, filters=None, *, actor_key="warehouse_user"):
    dataset = build_report_dataset(
        actor=context[actor_key],
        company=context["company"],
        report_key=report_key,
        filters=filters or {},
    )
    return dataset, list(dataset.rows)


def test_dashboard_source_reports_low_stock_stock_and_unit_grouping():
    context = report_context()
    dataset, stock = rows(context, "supply_stock_balance")
    assert dataset.row_count == 2
    paper = next(row for row in stock if row["item_code"] == "S18PAPER")
    assert paper["current_quantity"] == Decimal("7.0000")
    assert paper["current_amount"] == Decimal("700.00")
    assert paper["is_low_stock"] is True
    assert paper["shortage_quantity"] == Decimal("3.0000")

    _, low = rows(context, "supply_low_stock", {"low_stock_scope": "formal"})
    assert len(low) == 1
    assert low[0]["item_code"] == "S18PAPER"
    assert low[0]["default_warehouse"] == context["warehouse"].name


def test_stock_movement_and_issue_summaries_reconcile_exactly():
    context = report_context()
    period = {"date_from": date(2026, 8, 1), "date_to": date(2026, 8, 31)}
    _, movement = rows(context, "supply_stock_movement", period)
    paper = next(row for row in movement if row["item_code"] == "S18PAPER")
    assert paper["opening_quantity"] == Decimal("0.0000")
    assert paper["receipt_quantity"] == Decimal("12.0000")
    assert paper["issue_quantity"] == Decimal("5.0000")
    assert paper["ending_quantity"] == Decimal("7.0000")
    assert paper["receipt_amount"] - paper["issue_amount"] == paper["ending_amount"]

    _, details = rows(context, "supply_issue_detail", period)
    assert {row["item"].split(" / ")[0] for row in details} == {
        "S18PAPER",
        "S18CHAIR",
    }
    paper_detail = next(row for row in details if row["item"].startswith("S18PAPER"))
    assert paper_detail["current_net_quantity"] == Decimal("5.0000")
    assert paper_detail["current_net_amount"] == Decimal("500.00")

    _, department = rows(context, "supply_department_issue", period)
    assert sum((row["net_quantity"] for row in department), Decimal("0")) == Decimal("6.0000")
    _, employee = rows(context, "supply_employee_issue", period)
    assert {row["employee"] for row in employee} == {context["employee"].name}


def test_cross_period_reversal_is_negative_original_business_bucket():
    context = report_context()
    reverse_supply_document(
        document=context["paper_issue"],
        actor=context["warehouse_user"],
        reason="报表跨期间冲销测试",
        idempotency_key="s18-paper-reversal",
    )
    reversal_day = timezone.localdate()
    _, movement = rows(
        context,
        "supply_stock_movement",
        {"date_from": reversal_day, "date_to": reversal_day},
    )
    paper = next(row for row in movement if row["item_code"] == "S18PAPER")
    assert paper["opening_quantity"] == Decimal("7.0000")
    assert paper["issue_quantity"] == Decimal("-5.0000")
    assert paper["ending_quantity"] == Decimal("12.0000")
    assert (
        paper["opening_quantity"]
        + paper["receipt_quantity"]
        + paper["return_quantity"]
        + paper["transfer_in_quantity"]
        + paper["count_gain_quantity"]
        - paper["issue_quantity"]
        - paper["transfer_out_quantity"]
        - paper["count_loss_quantity"]
    ) == paper["ending_quantity"]
    _, details = rows(
        context,
        "supply_issue_detail",
        {
            "date_from": context["paper_issue"].business_date,
            "date_to": reversal_day,
        },
    )
    paper_rows = [row for row in details if row["item"].startswith("S18PAPER")]
    assert {row["business_type"] for row in paper_rows} == {"领用", "冲销"}
    assert all(row["current_net_quantity"] == Decimal("0.0000") for row in paper_rows)


def test_custody_reports_and_management_amount_keep_sources_separate():
    context = report_context()
    _, custodies = rows(context, "supply_custody_balance")
    assert len(custodies) == 1
    assert custodies[0]["current_quantity"] == Decimal("1.0000")
    assert custodies[0]["root_source_type"] == "领用"
    assert custodies[0]["current_amount"] == Decimal("80.00")

    _, movements = rows(context, "supply_custody_movement")
    assert len(movements) == 1
    assert movements[0]["action"] == "领用建立"

    _, amounts = rows(context, "supply_management_amount")
    labels = {row["component"] for row in amounts}
    assert "数量型耐用品仓库库存" in labels
    assert "数量型耐用品开放保管" in labels
    assert "数量型耐用品管理金额小计" in labels
    controlled = next(row for row in amounts if row["component"] == "逐件受控非固定资产")
    assert controlled["quantity"] == Decimal("0.0000")
    assert controlled.get("supply_amount") is None


def test_count_difference_report_covers_warehouse_snapshot_and_difference():
    context = report_context()
    task = create_supply_count_task(
        actor=context["warehouse_user"],
        company=context["company"],
        data={
            "name": "Sprint 18 报表盘点",
            "count_domain": "warehouse_stock",
            "warehouse": context["warehouse"],
            "planned_start": date(2026, 8, 27),
            "planned_end": date(2026, 8, 28),
            "idempotency_key": "s18-report-count",
        },
    )
    publish_supply_count_task(task=task, actor=context["warehouse_user"])
    line = task.lines.get(item=context["paper"])
    record_supply_count(
        line=line,
        counted_quantity=Decimal("6.0000"),
        remark="现场少一箱",
        actor=context["warehouse_user"],
    )
    _, result = rows(
        context,
        "supply_count_difference",
        {"differences_only": True, "count_domain": "warehouse_stock"},
    )
    assert len(result) == 1
    assert result[0]["task_no"] == task.task_no
    assert result[0]["difference_quantity"] == Decimal("-1.0000")
    assert result[0]["expected_amount"] == Decimal("700.00")


def test_controlled_non_fixed_report_uses_assets_and_asset_finance_permission():
    context, asset, _qr = formal_asset_context(
        "S18CNF", cost=Decimal("2000.00")
    )
    finance = build_report_dataset(
        actor=context["finance"],
        company=context["company"],
        report_key="controlled_non_fixed_assets",
        filters={},
    )
    finance_rows = list(finance.rows)
    assert len(finance_rows) == 1
    assert finance_rows[0]["asset_code"] == asset.asset_code
    assert finance_rows[0]["original_cost"] == Decimal("2000.00")
    assert "actual_book_value" not in finance_rows[0]

    admin = build_report_dataset(
        actor=context["admin"],
        company=context["company"],
        report_key="controlled_non_fixed_assets",
        filters={},
    )
    assert "original_cost" not in {
        column.key for column in admin.definition.columns
    }
    assert "original_cost" not in list(admin.rows)[0]


def test_employee_scope_excludes_cost_projection_and_company_stock():
    context = report_context()
    period = {"date_from": date(2026, 8, 1), "date_to": date(2026, 8, 31)}
    with pytest.raises(PermissionDenied):
        rows(context, "supply_stock_balance", actor_key="employee_user")
    with CaptureQueriesContext(connection) as captured:
        dataset, details = rows(
            context,
            "supply_issue_detail",
            period,
            actor_key="employee_user",
        )
    keys = {column.key for column in dataset.definition.columns}
    assert "amount" not in keys
    assert "unit_cost" not in keys
    assert "current_net_amount" not in keys
    assert all("amount" not in row and "unit_cost" not in row for row in details)
    data_selects = "\n".join(
        query["sql"]
        for query in captured.captured_queries
        if "supplies_supplystockledger" in query["sql"]
    ).lower()
    assert '"amount_delta"' not in data_selects
    assert '"posted_amount"' not in data_selects
    assert '"posted_unit_cost"' not in data_selects

    output = BytesIO()
    write_report_workbook(dataset, output)
    output.seek(0)
    workbook = load_workbook(output, read_only=True, data_only=False)
    sheet = workbook[dataset.definition.sheet_name]
    headers = [cell.value for cell in next(sheet.iter_rows(max_row=1))]
    assert "金额" not in headers
    assert "单位成本" not in headers
    assert "当前净领用金额" not in headers


def test_excel_keeps_date_quantity_cost_and_amount_as_native_cells():
    context = report_context()
    dataset, _ = rows(context, "supply_stock_ledger")
    output = BytesIO()
    write_report_workbook(dataset, output)
    output.seek(0)
    workbook = load_workbook(output, read_only=False, data_only=False)
    sheet = workbook[dataset.definition.sheet_name]
    headers = [cell.value for cell in sheet[1]]
    first = {header: sheet.cell(2, index + 1) for index, header in enumerate(headers)}
    assert first["业务日期"].is_date
    assert first["数量变动"].data_type == "n"
    assert first["金额变动"].data_type == "n"
    assert first["过账单位成本"].data_type == "n"


def test_export_log_records_actual_rows_filters_and_cost_scope(settings, tmp_path):
    context = report_context()
    media_root = tmp_path / "media"
    temp_root = tmp_path / "tmp"
    media_root.mkdir()
    temp_root.mkdir()
    settings.MEDIA_ROOT = media_root
    settings.IMPORT_TEMP_ROOT = temp_root
    from django.core.files.storage import storages

    storages._storages.clear()
    export_log = generate_report_export(
        actor=context["warehouse_user"],
        company=context["company"],
        report_key="supply_stock_balance",
        filters={},
        idempotency_key="s18-stock-export",
    )
    export_log.refresh_from_db()
    assert export_log.status == ExportLog.Status.COMPLETED
    assert export_log.row_count == 2
    assert export_log.filters_json["_includes_cost_fields"] is True
    assert "current_amount" in export_log.filters_json["_cost_columns"]
    assert export_log.output_attachment.sha256 == export_log.output_sha256

    employee_export = generate_report_export(
        actor=context["employee_user"],
        company=context["company"],
        report_key="supply_issue_detail",
        filters={"date_from": date(2026, 8, 1), "date_to": date(2026, 8, 31)},
        idempotency_key="s18-employee-export",
    )
    employee_export.refresh_from_db()
    assert employee_export.row_count == 2
    assert employee_export.filters_json["_includes_cost_fields"] is False
    assert employee_export.filters_json["_cost_columns"] == []


def test_all_twelve_supply_reports_generate_real_xlsx_workbooks():
    context = report_context()
    period = {"date_from": date(2026, 8, 1), "date_to": date(2026, 8, 31)}
    report_filters = {
        "supply_stock_balance": {},
        "supply_low_stock": {"low_stock_scope": "formal"},
        "supply_stock_movement": period,
        "supply_stock_ledger": period,
        "supply_issue_detail": period,
        "supply_department_issue": period,
        "supply_employee_issue": period,
        "supply_custody_balance": {},
        "supply_custody_movement": period,
        "supply_count_difference": {},
        "controlled_non_fixed_assets": {},
        "supply_management_amount": {},
    }
    generated = 0
    for report_key, filters in report_filters.items():
        dataset = build_report_dataset(
            actor=context["warehouse_user"],
            company=context["company"],
            report_key=report_key,
            filters=filters,
        )
        output = BytesIO()
        write_report_workbook(dataset, output)
        output.seek(0)
        workbook = load_workbook(output, read_only=True, data_only=False)
        assert "导出说明" in workbook.sheetnames
        assert dataset.definition.sheet_name in workbook.sheetnames
        generated += 1
    assert generated == 12


def test_completed_export_is_reauthorized_against_cost_columns(settings, tmp_path):
    context, _asset, _qr = formal_asset_context(
        "S18EXPAUTH", cost=Decimal("2000.00")
    )
    media_root = tmp_path / "media"
    temp_root = tmp_path / "tmp"
    media_root.mkdir()
    temp_root.mkdir()
    settings.MEDIA_ROOT = media_root
    settings.IMPORT_TEMP_ROOT = temp_root
    from django.core.files.storage import storages

    storages._storages.clear()
    export_log = generate_report_export(
        actor=context["finance"],
        company=context["company"],
        report_key="controlled_non_fixed_assets",
        filters={},
        idempotency_key="s18-controlled-export-auth",
    )
    export_log.refresh_from_db()
    assert export_log.filters_json["_cost_columns"] == ["original_cost"]
    assert can_download_export(context["finance"], export_log) is True
    assert can_download_export(context["admin"], export_log) is False
    assert can_download_export(context["equipment"], export_log) is False


def test_supply_report_http_permissions_pagination_and_dashboard_drilldown(client):
    context = report_context()
    client.force_login(context["warehouse_user"])
    dashboard = client.get(reverse("supplies:dashboard"))
    assert dashboard.status_code == 200
    assert dashboard.context["dashboard"]["low_stock_count"] == 1
    assert reverse("reports:supply-report-index").encode() in dashboard.content
    report = client.get(
        reverse("reports:supply-report-detail", args=["supply_low_stock"]),
        {"low_stock_scope": "formal"},
    )
    assert report.status_code == 200
    assert report.context["dataset"].row_count == 1

    client.force_login(context["employee_user"])
    denied = client.get(
        reverse("reports:supply-report-detail", args=["supply_stock_balance"])
    )
    assert denied.status_code == 403
    allowed = client.get(
        reverse("reports:supply-report-detail", args=["supply_issue_detail"]),
        {"date_from": "2026-08-01", "date_to": "2026-08-31"},
    )
    assert allowed.status_code == 200
    assert b"\xe9\x87\x91\xe9\xa2\x9d" not in allowed.content


def test_department_manager_management_and_unassigned_role_report_boundaries(client):
    context = report_context()
    manager = make_user("s18-report-manager", "department_manager")
    UserDepartmentScope.objects.create(
        company=context["company"],
        user=manager,
        department=context["department"],
        include_descendants=True,
        is_active=True,
        assigned_by=context["warehouse_user"],
    )
    client.force_login(manager)
    issue = client.get(
        reverse("reports:supply-report-detail", args=["supply_issue_detail"]),
        {"date_from": "2026-08-01", "date_to": "2026-08-31"},
    )
    assert issue.status_code == 200
    assert issue.context["dataset"].row_count == 2
    assert "amount" not in {
        column.key for column in issue.context["dataset"].definition.columns
    }
    assert client.get(
        reverse("reports:supply-report-detail", args=["supply_stock_ledger"])
    ).status_code == 403
    assert client.post(
        reverse("reports:supply-report-export", args=["supply_stock_balance"]),
        {"idempotency_key": "denied-stock-export"},
    ).status_code == 403

    management = make_user("s18-report-management", "management")
    client.force_login(management)
    stock = client.get(
        reverse("reports:supply-report-detail", args=["supply_stock_balance"])
    )
    assert stock.status_code == 200
    assert "current_amount" in {
        column.key for column in stock.context["dataset"].definition.columns
    }
    dashboard = client.get(reverse("supplies:dashboard"))
    assert dashboard.status_code == 200
    assert reverse("supplies:item-create").encode() not in dashboard.content

    unassigned = make_user("s18-report-unassigned")
    client.force_login(unassigned)
    assert client.get(reverse("supplies:dashboard")).status_code == 403
    assert client.get(reverse("reports:supply-report-index")).status_code == 403
