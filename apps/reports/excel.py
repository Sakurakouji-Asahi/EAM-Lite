"""Excel writers for materialized report DTOs; this module performs no DB I/O."""

from __future__ import annotations

from datetime import date, datetime
from decimal import Decimal

from django.utils import timezone
from openpyxl import Workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from apps.reports.schemas import CellKind, TPLUS_ENTRY_COLUMNS


XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
FORMULA_PREFIXES = ("=", "+", "-", "@")
HEADER_FILL = PatternFill("solid", fgColor="D9EAF7")
HEADER_FONT = Font(bold=True)
MONEY_FORMAT = '#,##0.00;[Red]-#,##0.00'
RATE_FORMAT = "0.00%"
DATE_FORMAT = "yyyy-mm-dd"
DATETIME_FORMAT = "yyyy-mm-dd hh:mm:ss"


def _safe_text(value):
    text = "" if value is None else str(value)
    if text.startswith(FORMULA_PREFIXES):
        return "'" + text
    return text


def _excel_datetime(value):
    if isinstance(value, datetime) and timezone.is_aware(value):
        return timezone.localtime(value).replace(tzinfo=None)
    return value


def _metadata_kind(value):
    if isinstance(value, bool):
        return CellKind.BOOLEAN
    if isinstance(value, datetime):
        return CellKind.DATETIME
    if isinstance(value, date):
        return CellKind.DATE
    if isinstance(value, Decimal):
        return CellKind.DECIMAL
    if isinstance(value, int):
        return CellKind.INTEGER
    return CellKind.TEXT


def _cell(ws, value, kind=CellKind.TEXT, *, header=False):
    if header:
        cell = WriteOnlyCell(ws, value=_safe_text(value))
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        return cell
    if kind in {CellKind.TEXT, CellKind.IDENTIFIER}:
        cell = WriteOnlyCell(ws, value=_safe_text(value))
        if kind == CellKind.IDENTIFIER:
            cell.number_format = "@"
        return cell
    cell = WriteOnlyCell(ws, value=_excel_datetime(value))
    if kind in {CellKind.MONEY, CellKind.DECIMAL}:
        cell.number_format = MONEY_FORMAT
    elif kind == CellKind.RATE:
        cell.number_format = RATE_FORMAT
    elif kind == CellKind.DATE:
        cell.number_format = DATE_FORMAT
    elif kind == CellKind.DATETIME:
        cell.number_format = DATETIME_FORMAT
    return cell


def _configure_table(ws, columns, row_count):
    ws.freeze_panes = "A2"
    last = get_column_letter(len(columns))
    ws.auto_filter.ref = f"A1:{last}{max(1, row_count + 1)}"
    for index, column in enumerate(columns, start=1):
        ws.column_dimensions[get_column_letter(index)].width = column.width


def _append_table(ws, columns, rows):
    _configure_table(ws, columns, len(rows))
    ws.append([_cell(ws, column.label, header=True) for column in columns])
    for row in rows:
        ws.append([_cell(ws, row.get(column.key), column.kind) for column in columns])


def write_report_workbook(dataset, destination, *, generated_at=None):
    """Write one fixed-schema generic report workbook to a path/file object."""
    generated_at = generated_at or timezone.now()
    workbook = Workbook(write_only=True)
    workbook.properties.creator = "EAM-Lite"
    workbook.properties.created = _excel_datetime(generated_at)
    information = workbook.create_sheet("导出说明")
    information.column_dimensions["A"].width = 24
    information.column_dimensions["B"].width = 72
    information_rows = [
        ("报表类型", dataset.definition.title),
        ("生成完成时间", generated_at),
        ("数据截止时间", dataset.data_snapshot_at),
        ("Schema版本", dataset.definition.schema_version),
        ("数据行数", dataset.row_count),
    ]
    information_rows.extend(
        (f"筛选条件：{key}", value)
        for key, value in dataset.filters.items() if not key.startswith("_")
    )
    for label, value in information_rows:
        kind = _metadata_kind(value)
        information.append([_cell(information, label, header=True), _cell(information, value, kind)])
    information.protection.sheet = True
    sheet = workbook.create_sheet(dataset.definition.sheet_name[:31])
    _append_table(sheet, dataset.definition.columns, dataset.rows)
    workbook.save(destination)


def _write_tplus_information(
    workbook, *, dataset, export_id, company_name, requested_by, generated_at
):
    ws = workbook.create_sheet("导出说明")
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 72
    rows = (
        ("用途", "仅供T+人工对账，不代表已记账"),
        ("公司", company_name),
        ("会计期间", dataset.filters.get("period", "")),
        ("数据截止时间", dataset.data_snapshot_at),
        ("生成完成时间", generated_at),
        ("导出人", requested_by),
        ("导出编号", export_id),
        ("Schema版本", dataset.definition.schema_version),
        ("数据口径", "仅使用已确认财务数据和已过账实际分录；理论试算不进入实际金额。"),
        ("累计折旧勾稽", "期末累计折旧=期初累计折旧+本期自动折旧+本期手工折旧+本期调整净额+本期冲销净额"),
        ("净值勾稽", "期末账面净值=原值-期末累计折旧-减值准备"),
        ("匹配优先级", "T+资产卡片编码精确匹配；为空时按EAM资产编码精确匹配；禁止模糊匹配。"),
        ("粘贴区容量", "最多5000行；粘贴T+数据后查看“对账差异”。"),
    )
    for label, value in rows:
        kind = _metadata_kind(value)
        ws.append([_cell(ws, label, header=True), _cell(ws, value, kind)])
    for key, value in dataset.filters.items():
        if key.startswith("_"):
            continue
        kind = _metadata_kind(value)
        ws.append([_cell(ws, f"筛选条件：{key}", header=True), _cell(ws, value, kind)])
    ws.protection.sheet = True


def _write_tplus_paste_sheet(workbook):
    ws = workbook.create_sheet("T+数据粘贴区")
    labels = (
        "T+资产卡片编码", "EAM资产编号（如T+已维护）", "资产名称", "原值",
        "期初累计折旧", "本期折旧", "期末累计折旧", "减值准备", "期末账面净值",
    )
    widths = (22, 20, 24, 16, 18, 16, 16, 16, 28)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = "A1:I5001"
    for index, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(index)].width = width
    ws.append([_cell(ws, label, header=True) for label in labels])
    # Pre-create the approved 5,000-row input area so the identifier cells
    # themselves carry Excel's text format.  A column dimension style is not
    # inherited by empty cells after an xlsx round-trip, which would let a
    # pasted code such as 000123 lose its leading zeroes.
    for _ in range(5000):
        card = WriteOnlyCell(ws, value=None)
        card.number_format = "@"
        asset_code = WriteOnlyCell(ws, value=None)
        asset_code.number_format = "@"
        ws.append([card, asset_code, None, None, None, None, None, None, None])
    return ws


def _lookup_formula(card_cell, code_cell, value_column, eam_last_row):
    paste = "'T+数据粘贴区'"
    return (
        f'=IF({card_cell}<>"",IF(COUNTIF({paste}!$A$2:$A$5001,{card_cell})>1,"",'
        f'IF(COUNTIF({paste}!$A$2:$A$5001,{card_cell})=1,'
        f'INDEX({paste}!${value_column}$2:${value_column}$5001,MATCH({card_cell},{paste}!$A$2:$A$5001,0)),'
        f'IF(COUNTIF({paste}!$B$2:$B$5001,{code_cell})=1,'
        f'INDEX({paste}!${value_column}$2:${value_column}$5001,MATCH({code_cell},{paste}!$B$2:$B$5001,0)),""))),'
        f'IF(COUNTIF({paste}!$B$2:$B$5001,{code_cell})=1,'
        f'INDEX({paste}!${value_column}$2:${value_column}$5001,MATCH({code_cell},{paste}!$B$2:$B$5001,0)),""))'
    )


def _match_formula(row_number):
    paste = "'T+数据粘贴区'"
    card, code = f"C{row_number}", f"B{row_number}"
    return (
        f'=IF({card}<>"",IF(COUNTIF({paste}!$A$2:$A$5001,{card})>1,"duplicate key",'
        f'IF(COUNTIF({paste}!$A$2:$A$5001,{card})=1,"matched",'
        f'IF(COUNTIF({paste}!$B$2:$B$5001,{code})>1,"duplicate key",'
        f'IF(COUNTIF({paste}!$B$2:$B$5001,{code})=1,"matched","EAM only")))),'
        f'IF(COUNTIF({paste}!$B$2:$B$5001,{code})>1,"duplicate key",'
        f'IF(COUNTIF({paste}!$B$2:$B$5001,{code})=1,"matched","EAM only")))'
    )


def _write_tplus_difference_sheet(workbook, dataset):
    ws = workbook.create_sheet("对账差异")
    labels = (
        "匹配状态", "EAM资产编码", "T+资产卡片编码", "资产名称",
        "EAM原值", "T+原值", "原值差异",
        "EAM累计折旧", "T+累计折旧", "累计折旧差异",
        "EAM减值准备", "T+减值准备", "减值准备差异",
        "EAM账面净值", "T+账面净值", "账面净值差异",
        "EAM本期折旧", "T+本期折旧", "本期折旧差异", "差异原因", "处理备注",
    )
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:U{max(2, len(dataset.asset_rows) + 5001)}"
    for index in range(1, len(labels) + 1):
        ws.column_dimensions[get_column_letter(index)].width = 18
    ws.column_dimensions["D"].width = 24
    ws.column_dimensions["T"].width = 24
    ws.column_dimensions["U"].width = 28
    ws.append([_cell(ws, label, header=True) for label in labels])

    for row_number, row in enumerate(dataset.asset_rows, start=2):
        current_depreciation = sum(
            (row[key] for key in ("automatic_depreciation", "manual_depreciation", "adjustment_net", "reversal_net")),
            Decimal("0.00"),
        )
        values = [
            _match_formula(row_number), row["asset_code"], row["tplus_card_code"], row["asset_name"],
            row["original_cost"], _lookup_formula(f"C{row_number}", f"B{row_number}", "D", len(dataset.asset_rows) + 1), f'=IF(A{row_number}="matched",E{row_number}-F{row_number},"")',
            row["ending_accumulated_depreciation"], _lookup_formula(f"C{row_number}", f"B{row_number}", "G", len(dataset.asset_rows) + 1), f'=IF(A{row_number}="matched",H{row_number}-I{row_number},"")',
            row["impairment"], _lookup_formula(f"C{row_number}", f"B{row_number}", "H", len(dataset.asset_rows) + 1), f'=IF(A{row_number}="matched",K{row_number}-L{row_number},"")',
            row["ending_book_value"], _lookup_formula(f"C{row_number}", f"B{row_number}", "I", len(dataset.asset_rows) + 1), f'=IF(A{row_number}="matched",N{row_number}-O{row_number},"")',
            current_depreciation, _lookup_formula(f"C{row_number}", f"B{row_number}", "F", len(dataset.asset_rows) + 1), f'=IF(A{row_number}="matched",Q{row_number}-R{row_number},"")',
            "", "",
        ]
        formula_columns = {1, 6, 7, 9, 10, 12, 13, 15, 16, 18, 19}
        cells = []
        for index, value in enumerate(values, start=1):
            kind = CellKind.IDENTIFIER if index in (2, 3) else CellKind.MONEY if index in range(5, 20) else CellKind.TEXT
            cell = WriteOnlyCell(ws, value=value if index in formula_columns else _safe_text(value) if kind in {CellKind.TEXT, CellKind.IDENTIFIER} else value)
            if kind == CellKind.IDENTIFIER:
                cell.number_format = "@"
            elif kind == CellKind.MONEY:
                cell.number_format = MONEY_FORMAT
            cells.append(cell)
        ws.append(cells)

    first_extra = len(dataset.asset_rows) + 2
    for paste_row in range(2, 5002):
        output_row = first_extra + paste_row - 2
        paste = "'T+数据粘贴区'"
        card, code = f"{paste}!A{paste_row}", f"{paste}!B{paste_row}"
        status = (
            f'=IF(COUNTA({paste}!A{paste_row}:B{paste_row})=0,"",'
            f'IF({card}<>"",'
            f'IF(OR(COUNTIF({paste}!$A$2:$A$5001,{card})>1,'
            f'COUNTIF(\'EAM固定资产明细\'!$B$2:$B$1048576,{card})>1),'
            f'"duplicate key",IF(COUNTIF(\'EAM固定资产明细\'!$B$2:$B$1048576,{card})=1,"",'
            f'IF({code}<>"",IF(OR(COUNTIF({paste}!$B$2:$B$5001,{code})>1,'
            f'COUNTIF(\'EAM固定资产明细\'!$A$2:$A$1048576,{code})>1),"duplicate key",'
            f'IF(COUNTIF(\'EAM固定资产明细\'!$A$2:$A$1048576,{code})=1,"","T+ only")),"T+ only"))),'
            f'IF(OR(COUNTIF({paste}!$B$2:$B$5001,{code})>1,'
            f'COUNTIF(\'EAM固定资产明细\'!$A$2:$A$1048576,{code})>1),'
            f'"duplicate key",IF(COUNTIF(\'EAM固定资产明细\'!$A$2:$A$1048576,{code})=1,"","T+ only"))))'
        )
        formulas = [
            status, f"={code}", f"={card}", f"={paste}!C{paste_row}", "", f"={paste}!D{paste_row}", f'=IF(A{output_row}="T+ only",-F{output_row},"")',
            "", f"={paste}!G{paste_row}", f'=IF(A{output_row}="T+ only",-I{output_row},"")',
            "", f"={paste}!H{paste_row}", f'=IF(A{output_row}="T+ only",-L{output_row},"")',
            "", f"={paste}!I{paste_row}", f'=IF(A{output_row}="T+ only",-O{output_row},"")',
            "", f"={paste}!F{paste_row}", f'=IF(A{output_row}="T+ only",-R{output_row},"")', "", "",
        ]
        ws.append([WriteOnlyCell(ws, value=value) for value in formulas])


def write_tplus_workbook(
    dataset, destination, *, export_id, company_name, requested_by, generated_at
):
    """Write the fixed five-sheet T+ reconciliation workbook."""
    workbook = Workbook(write_only=True)
    workbook.properties.creator = "EAM-Lite"
    workbook.properties.created = _excel_datetime(generated_at)
    _write_tplus_information(
        workbook, dataset=dataset, export_id=str(export_id), company_name=company_name,
        requested_by=requested_by, generated_at=generated_at,
    )
    assets = workbook.create_sheet("EAM固定资产明细")
    _append_table(assets, dataset.definition.columns, dataset.asset_rows)
    assets.protection.sheet = True
    entries = workbook.create_sheet("本期折旧分录")
    _append_table(entries, TPLUS_ENTRY_COLUMNS, dataset.entry_rows)
    entries.protection.sheet = True
    _write_tplus_paste_sheet(workbook)
    _write_tplus_difference_sheet(workbook, dataset)
    workbook.save(destination)


__all__ = ["XLSX_MIME", "write_report_workbook", "write_tplus_workbook"]
