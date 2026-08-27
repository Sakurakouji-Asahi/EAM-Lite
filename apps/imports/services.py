"""Versioned, protected XLSX imports for Sprint 1 master data.

The policy is deliberately conservative: imports only create records.  They
never update by code or name, and any existing unique key is a row error.
"""

from __future__ import annotations

import hashlib
import io
import re
import uuid
import zipfile
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from decimal import Decimal, InvalidOperation

from defusedxml import ElementTree as DefusedElementTree

from django.core.exceptions import PermissionDenied, ValidationError
from django.core.files.base import ContentFile
from django.core.files.storage import default_storage
from django.db import connection, transaction
from django.db.models import Q
from django.utils import timezone
from django.utils.text import get_valid_filename
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation

from apps.audit.services import request_audit_context, write_business_audit_log
from apps.assets.permissions import can_create_asset_draft
from apps.assets.services import _validate_filename, create_asset_draft
from apps.finance.permissions import can_manage_finance
from apps.masterdata.normalization import clean_display_identifier, normalize_identifier
from apps.masterdata.permissions import (
    current_company,
    role_names_for,
    require_manage_masterdata,
)
from apps.masterdata.services import (
    create_department,
    create_employee,
    get_system_setting,
)
from apps.imports.tempfiles import hold_temp_file_active


XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
MAX_ARCHIVE_MEMBERS = 512
MAX_UNCOMPRESSED_BYTES = 20 * 1024 * 1024
MAX_ARCHIVE_MEMBER_BYTES = 10 * 1024 * 1024
MAX_COMPRESSION_RATIO = 100
MAX_WORKSHEET_ROWS = 10_001
MAX_WORKSHEET_COLUMNS = 128
MAX_WORKSHEET_CELLS = MAX_WORKSHEET_ROWS * MAX_WORKSHEET_COLUMNS
MAX_IMPORT_ROWS = 10_000
CELL_REFERENCE_PATTERN = re.compile(r"^\$?([A-Za-z]{1,3})\$?([1-9][0-9]*)$")
FORBIDDEN_ARCHIVE_PARTS = (
    "vbaproject.bin",
    "xl/externallinks/",
    "xl/embeddings/",
    "xl/activex/",
)
FORBIDDEN_RELATIONSHIP_TYPES = (
    "/externallink",
    "/oleobject",
    "/controlproperties",
    "/activex",
    "/relationships/package",
    "/attachedtemplate",
)


def _lock_import_namespace(company_id):
    """Serialize duplicate/idempotency decisions per company on PostgreSQL."""
    from django.db import connection

    if connection.vendor == "postgresql":
        key = int(company_id)
        if not -(2**31) <= key < 2**31:
            raise ValidationError("公司标识超出导入锁支持范围。")
        with connection.cursor() as cursor:
            cursor.execute(
                "SELECT pg_advisory_xact_lock(%s, %s)",
                [0x45414D49, key],
            )


@dataclass(frozen=True)
class Column:
    name: str
    key: str
    required: bool = False


@dataclass(frozen=True)
class TemplateDefinition:
    import_type: str
    label: str
    version: str
    sheet_name: str
    columns: tuple[Column, ...]
    has_example_sheet: bool = False

    @property
    def headers(self):
        return tuple(column.name for column in self.columns)


TEMPLATE_REGISTRY = {
    "department": TemplateDefinition(
        import_type="department",
        label="部门",
        version="department-v1",
        sheet_name="部门导入",
        columns=(
            Column("部门编码", "code", True),
            Column("部门名称", "name", True),
            Column("上级部门编码", "parent_code"),
            Column("经理工号", "manager_employee_no"),
            Column("是否启用", "is_active"),
        ),
    ),
    "employee": TemplateDefinition(
        import_type="employee",
        label="人员",
        version="employee-v1",
        sheet_name="人员导入",
        columns=(
            Column("员工编号", "employee_no", True),
            Column("姓名", "name", True),
            Column("部门编码", "department_code", True),
            Column("任职状态", "employment_status", True),
            Column("入职日期", "hire_date"),
            Column("离职日期", "termination_date"),
            Column("手机号码", "mobile"),
            Column("备注", "remark"),
            Column("是否启用", "is_active"),
        ),
    ),
    "asset_initialization": TemplateDefinition(
        import_type="asset_initialization",
        label="资产初始化",
        version="asset-initialization-v2",
        sheet_name="资产初始化导入",
        has_example_sheet=True,
        columns=(
            Column("资产名称", "asset_name", True),
            Column("实物分类编码", "category_code", True),
            Column("品牌", "brand"),
            Column("型号", "model"),
            Column("厂家", "manufacturer"),
            Column("序列号", "serial_number"),
            Column("出厂编号", "factory_number"),
            Column("历史参考编号", "historical_code"),
            Column("数量", "quantity", True),
            Column("单位", "unit", True),
            Column("公司编码", "company_code", True),
            Column("部门编码", "department_code", True),
            Column("责任员工编号", "responsible_employee_no", True),
            Column("位置编码", "location_code", True),
            Column("购置日期", "acquisition_date"),
            Column("达到可使用状态日期", "commissioning_date"),
            Column("是否需要保养", "is_maintenance_required"),
            Column("附件后续上传说明", "attachment_note"),
            Column("备注", "notes"),
            Column("会计认定", "accounting_treatment"),
            Column("会计认定说明", "accounting_treatment_reason"),
            Column("固定资产类别编码", "fixed_asset_category_code"),
            Column("原值", "original_cost"),
            Column("资本化日期", "capitalization_date"),
            Column("折旧政策编码", "depreciation_policy_key"),
            Column("折旧方法", "method"),
            Column("计提周期", "posting_period"),
            Column("起算规则", "start_rule"),
            Column("指定起算日期", "specified_start"),
            Column("历史起算原因", "historical_start_reason"),
            Column("停止规则", "stop_rule"),
            Column("使用寿命月数", "useful_life_months"),
            Column("残值方式", "salvage_mode"),
            Column("残值率", "salvage_rate"),
            Column("残值金额", "salvage_amount"),
            Column("年度计提月份", "annual_posting_month"),
            Column("预计总工作量", "expected_total_units"),
            Column("工作量单位", "work_unit"),
            Column("实际期初累计折旧", "opening_actual_accumulated_depreciation"),
            Column("期初减值", "opening_impairment"),
            Column("实际期初账面净值", "opening_book_value"),
            Column("实际接续日", "actual_continuation_date"),
            Column("理论测算截止日", "theoretical_as_of_date"),
            Column("财务备注", "finance_remark"),
        ),
    ),
    "item_master": TemplateDefinition(
        import_type="item_master",
        label="低值物品档案",
        version="supply-item-master-v1",
        sheet_name="低值物品档案导入",
        has_example_sheet=True,
        columns=(
            Column("物品编码", "item_code", True),
            Column("物品名称", "name", True),
            Column("分类编码", "category_code", True),
            Column("管理模式", "item_type", True),
            Column("单位", "unit", True),
            Column("规格", "specification"),
            Column("型号", "model"),
            Column("品牌", "brand"),
            Column("最低库存", "minimum_stock_quantity"),
            Column("默认仓库编码", "default_warehouse_code"),
            Column("备注", "remark"),
        ),
    ),
    "opening_stock": TemplateDefinition(
        import_type="opening_stock",
        label="低值物品期初库存",
        version="supply-opening-stock-v1",
        sheet_name="期初库存导入",
        has_example_sheet=True,
        columns=(
            Column("公司编码", "company_code", True),
            Column("仓库编码", "warehouse_code", True),
            Column("物品编码", "item_code", True),
            Column("数量", "quantity", True),
            Column("单位成本", "unit_cost", True),
            Column("0成本原因", "zero_cost_reason"),
            Column("备注", "remark"),
        ),
    ),
    "opening_custody": TemplateDefinition(
        import_type="opening_custody",
        label="耐用品期初保管",
        version="supply-opening-custody-v1",
        sheet_name="期初保管导入",
        has_example_sheet=True,
        columns=(
            Column("物品编码", "item_code", True),
            Column("责任部门编码", "department_code", True),
            Column("责任员工编号", "employee_no"),
            Column("数量", "quantity", True),
            Column("单位成本", "unit_cost", True),
            Column("开始日期", "started_on", True),
            Column("备注", "remark"),
        ),
    ),
}


def get_template_definition(import_type: str, *, company=None) -> TemplateDefinition:
    try:
        definition = TEMPLATE_REGISTRY[import_type]
    except KeyError as exc:
        raise ValidationError("不支持的导入类型。") from exc
    if import_type != "asset_initialization":
        return definition
    if company is None:
        return definition
    from apps.assets.models import AssetCustomField

    custom_columns = tuple(
        Column(f"自定义:{field.code}", f"custom:{field.normalized_code}")
        for field in AssetCustomField.objects.filter(
            company=company, is_active=True
        ).order_by("display_order", "normalized_code")
    )
    return TemplateDefinition(
        import_type=definition.import_type,
        label=definition.label,
        version=definition.version,
        sheet_name=definition.sheet_name,
        columns=definition.columns + custom_columns,
        has_example_sheet=True,
    )


def require_import_permission(
    actor, import_type, *, company=None, department=None
):
    """Apply the same import action gate to pages, downloads and Services."""

    if import_type == "asset_initialization":
        roles = role_names_for(actor)
        if roles.intersection({"finance", "equipment", "warehouse"}):
            return
        if "department_manager" in roles:
            if department is None:
                # Upload/template access is allowed; every staged row is still
                # re-authorized against its concrete department below.
                return
            if company is not None and can_create_asset_draft(
                actor, company, department
            ):
                return
        raise PermissionDenied("您没有导入资产初始化草稿的权限。")
    if import_type == "item_master":
        from apps.supplies.models import SupplyItemType
        from apps.supplies.permissions import require_manage_supply_item

        require_manage_supply_item(actor, SupplyItemType.DURABLE_QUANTITY)
        return
    if import_type == "opening_stock":
        from apps.supplies.permissions import require_create_supply_document

        require_create_supply_document(actor)
        return
    if import_type == "opening_custody":
        from apps.supplies.permissions import require_import_opening_custody

        require_import_opening_custody(actor)
        return
    resource = {"department": "department", "employee": "employee"}.get(import_type)
    if resource is None:
        raise ValidationError("不支持的导入类型。")
    require_manage_masterdata(actor, resource)


_require_import_permission = require_import_permission


def _require_current_import_company(company):
    current = current_company(include_inactive=True)
    if (
        company is None
        or current is None
        or getattr(company, "pk", None) != current.pk
    ):
        raise PermissionDenied("导入目标不属于当前公司。")


def build_template_workbook(import_type: str, company=None) -> bytes:
    definition = get_template_definition(import_type, company=company)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = definition.sheet_name
    sheet.append(definition.headers)
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="4472C4")
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:{sheet.cell(1, len(definition.columns)).coordinate}"
    for index, column in enumerate(definition.columns, 1):
        sheet.column_dimensions[sheet.cell(1, index).column_letter].width = max(
            14, len(column.name) * 2 + 4
        )
        sheet.cell(1, index).comment = None
    boolean_header = next(
        (
            header
            for header in ("是否启用", "是否需要保养")
            if header in definition.headers
        ),
        None,
    )
    if boolean_header:
        boolean_validation = DataValidation(type="list", formula1='"是,否"')
        sheet.add_data_validation(boolean_validation)
        boolean_column = definition.headers.index(boolean_header) + 1
        boolean_validation.add(
            f"{sheet.cell(2, boolean_column).coordinate}:"
            f"{sheet.cell(1001, boolean_column).coordinate}"
        )
    if import_type == "employee":
        status_validation = DataValidation(
            type="list", formula1='"active,leaving,resigned"'
        )
        sheet.add_data_validation(status_validation)
        status_column = definition.headers.index("任职状态") + 1
        status_validation.add(
            f"{sheet.cell(2, status_column).coordinate}:"
            f"{sheet.cell(1001, status_column).coordinate}"
        )
    if import_type == "asset_initialization":
        treatment_validation = DataValidation(
            type="list", formula1='"fixed_asset,controlled_non_fixed"'
        )
        sheet.add_data_validation(treatment_validation)
        treatment_column = definition.headers.index("会计认定") + 1
        treatment_validation.add(
            f"{sheet.cell(2, treatment_column).coordinate}:"
            f"{sheet.cell(1001, treatment_column).coordinate}"
        )
    if import_type == "item_master":
        item_type_validation = DataValidation(
            type="list", formula1='"consumable,durable_quantity"'
        )
        sheet.add_data_validation(item_type_validation)
        item_type_column = definition.headers.index("管理模式") + 1
        item_type_validation.add(
            f"{sheet.cell(2, item_type_column).coordinate}:"
            f"{sheet.cell(1001, item_type_column).coordinate}"
        )
        minimum_validation = DataValidation(
            type="decimal",
            operator="greaterThanOrEqual",
            formula1="0",
            allow_blank=True,
        )
        sheet.add_data_validation(minimum_validation)
        minimum_column = definition.headers.index("最低库存") + 1
        minimum_validation.add(
            f"{sheet.cell(2, minimum_column).coordinate}:"
            f"{sheet.cell(1001, minimum_column).coordinate}"
        )
    if import_type in {"opening_stock", "opening_custody"}:
        quantity_validation = DataValidation(
            type="decimal",
            operator="greaterThan",
            formula1="0",
            allow_blank=False,
        )
        cost_validation = DataValidation(
            type="decimal",
            operator="greaterThanOrEqual",
            formula1="0",
            allow_blank=False,
        )
        sheet.add_data_validation(quantity_validation)
        sheet.add_data_validation(cost_validation)
        quantity_column = definition.headers.index("数量") + 1
        cost_column = definition.headers.index("单位成本") + 1
        quantity_validation.add(
            f"{sheet.cell(2, quantity_column).coordinate}:"
            f"{sheet.cell(1001, quantity_column).coordinate}"
        )
        cost_validation.add(
            f"{sheet.cell(2, cost_column).coordinate}:"
            f"{sheet.cell(1001, cost_column).coordinate}"
        )
    instructions = workbook.create_sheet("填写说明")
    instructions.append(["项目", "说明"])
    policies = [
        ("模板版本", definition.version),
        ("工作表", definition.sheet_name),
        ("数据策略", "只新增，不按名称或编码覆盖现有记录"),
        ("日期格式", "YYYY-MM-DD"),
        ("布尔值", "是 / 否；留空默认为“是”"),
        ("空行", "忽略完全空白的数据行"),
        ("安全", "未知列、公式、外部链接、宏或嵌入对象会导致拒绝"),
    ]
    if import_type == "department":
        policies.append(
            ("匹配键", "上级部门按当前公司的部门编码匹配；可引用同文件新部门")
        )
        policies.append(("经理", "按当前公司已有员工编号匹配"))
    else:
        if import_type == "employee":
            policies.append(("匹配键", "部门按当前公司已有部门编码匹配"))
        elif import_type == "asset_initialization":
            policies.extend(
                (
                    ("匹配键", "公司/分类/部门/位置按 code，责任人按 employee_no 精确匹配"),
                    ("单件规则", "每一行只代表一件实物，数量必须精确为 1"),
                    ("确认结果", "仅创建草稿，不生成正式编号、二维码或实际折旧分录"),
                    ("附件", "本表只填写后续上传说明，不接受本机路径或 URL 自动抓取"),
                    ("财务列", "只有 finance 可提交；无财务权限时所有财务列必须留空"),
                    ("会计认定", "只能为 fixed_asset 或 controlled_non_fixed"),
                    ("折旧机器值", "方法/周期/起止/残值方式使用系统机器值，不接受别名"),
                    ("旧资产接续", "填写历史起算原因或期初累计折旧/减值时，必须另填实际接续日"),
                    ("动态字段", "自定义列按“自定义:<code>”精确匹配当前启用字段"),
                )
            )
        elif import_type == "item_master":
            policies.extend(
                (
                    ("匹配键", "分类和默认仓库按当前公司的规范化编码匹配"),
                    ("管理模式", "只能为 consumable 或 durable_quantity"),
                    ("数量精度", "最低库存最多 4 位小数，留空按 0 处理"),
                    ("确认结果", "整批创建物品档案，不创建库存余额或库存流水"),
                    ("逐件物品", "需要逐件二维码、序列号或单件责任时不要导入本表，改用资产模块"),
                )
            )
        elif import_type == "opening_stock":
            policies.extend(
                (
                    ("匹配键", "公司、仓库和物品按当前公司的规范化编码匹配"),
                    ("数量精度", "数量必须大于 0，最多 4 位小数"),
                    ("成本精度", "单位成本不得小于 0，最多 6 位小数"),
                    ("0 成本", "单位成本为 0 时必须填写明确的 0 成本原因"),
                    ("确认结果", "按仓库生成期初入库草稿，不自动过账或改变库存"),
                    ("过账", "确认导入后仍须进入每张期初单核对并执行过账"),
                )
            )
        elif import_type == "opening_custody":
            policies.extend(
                (
                    ("匹配键", "物品和部门按当前公司编码匹配，员工按员工编号匹配"),
                    ("物品模式", "只允许已启用 durable_quantity 数量型低值耐用品"),
                    ("责任员工", "可留空；填写时必须属于责任部门且在职、启用"),
                    ("数量精度", "数量必须大于 0，最多 4 位小数"),
                    ("成本精度", "单位成本不得小于 0，最多 6 位小数"),
                    ("确认结果", "直接建立开放保管及不可变 opening 流水，不改变仓库库存"),
                    ("确认后更正", "确认后不提供破坏性回滚；错误需后续受控更正"),
                )
            )
    for row in policies:
        instructions.append(row)
    instructions.column_dimensions["A"].width = 18
    instructions.column_dimensions["B"].width = 90
    if definition.has_example_sheet:
        example = workbook.create_sheet("示例")
        example.append(definition.headers)
        if import_type == "item_master":
            sample = {
                "物品编码": "PAPER-A4",
                "物品名称": "A4 复印纸（示例，请勿直接导入）",
                "分类编码": "OFFICE",
                "管理模式": "consumable",
                "单位": "箱",
                "最低库存": 5,
                "默认仓库编码": "OFFICE-WH",
            }
        elif import_type == "opening_stock":
            sample = {
                "公司编码": getattr(company, "code", "COMPANY"),
                "仓库编码": "OFFICE-WH",
                "物品编码": "PAPER-A4",
                "数量": 10,
                "单位成本": 100,
                "备注": "示例，请勿直接导入",
            }
        elif import_type == "opening_custody":
            sample = {
                "物品编码": "CHAIR-01",
                "责任部门编码": "DEPT",
                "责任员工编号": "E0001",
                "数量": 2,
                "单位成本": 80,
                "开始日期": timezone.localdate(),
                "备注": "示例，请勿直接导入",
            }
        else:
            sample = {
                "资产名称": "示例设备（请勿直接导入）",
                "实物分类编码": "EQUIPMENT",
                "数量": 1,
                "单位": "台",
                "公司编码": "COMPANY",
                "部门编码": "DEPT",
                "责任员工编号": "E0001",
                "位置编码": "POSITION-01",
                "是否需要保养": "是",
                "附件后续上传说明": "确认草稿后通过受保护附件入口上传照片",
            }
        example.append([sample.get(header, "") for header in definition.headers])
        example.freeze_panes = "A2"
    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


def _json_value(value):
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if value is None or isinstance(value, (str, int, float, bool)):
        return value
    return str(value)


ASSET_FINANCE_KEYS = frozenset(
    {
        "accounting_treatment",
        "accounting_treatment_reason",
        "fixed_asset_category_code",
        "original_cost",
        "capitalization_date",
        "depreciation_policy_key",
        "method",
        "posting_period",
        "start_rule",
        "specified_start",
        "historical_start_reason",
        "stop_rule",
        "useful_life_months",
        "salvage_mode",
        "salvage_rate",
        "salvage_amount",
        "annual_posting_month",
        "expected_total_units",
        "work_unit",
        "opening_actual_accumulated_depreciation",
        "opening_impairment",
        "opening_book_value",
        "actual_continuation_date",
        "theoretical_as_of_date",
        "finance_remark",
    }
)

def _decimal_value(value, field, errors, *, places, required=False, minimum=None):
    if value in (None, ""):
        if required:
            errors.append(_error(field, value, "必填。"))
        return None
    if isinstance(value, bool):
        errors.append(
            _error(field, value, "必须提供十进制数值，不能使用布尔值。")
        )
        return None
    try:
        result = value if isinstance(value, Decimal) else Decimal(str(value).strip())
    except (InvalidOperation, TypeError, ValueError):
        errors.append(_error(field, value, "不是有效的十进制数。"))
        return None
    if not result.is_finite():
        errors.append(_error(field, value, "必须是有限十进制数。"))
        return None
    exponent = -result.as_tuple().exponent
    if exponent > places:
        errors.append(_error(field, value, f"小数位不得超过 {places} 位。"))
        return None
    if minimum is not None and result < minimum:
        errors.append(_error(field, value, f"不得小于 {minimum}。"))
        return None
    return result


def _integer_value(value, field, errors, *, required=False, minimum=None, maximum=None):
    decimal_value = _decimal_value(
        value, field, errors, places=0, required=required, minimum=minimum
    )
    if decimal_value is None:
        return None
    integer = int(decimal_value)
    if maximum is not None and integer > maximum:
        errors.append(_error(field, value, f"不得大于 {maximum}。"))
        return None
    return integer


def _nullable_boolean(value, field, errors):
    if value in (None, ""):
        return None
    return _boolean(value, field, errors, default=False)


def _serialize_mapping(value):
    """Recursively make finance preview data safe for JSON staging."""

    if isinstance(value, Decimal):
        return str(value)
    if isinstance(value, (date, datetime)):
        return value.isoformat()
    if isinstance(value, dict):
        return {str(key): _serialize_mapping(item) for key, item in value.items()}
    if isinstance(value, (tuple, list)):
        return [_serialize_mapping(item) for item in value]
    return value


def _error(field, value, reason):
    return {"field": field, "value": _json_value(value), "reason": reason}


def _text(value):
    if value is None:
        return ""
    return clean_display_identifier(str(value))


def _boolean(value, field, errors, default=True):
    text = _text(value)
    if not text:
        return default
    if text == "是":
        return True
    if text == "否":
        return False
    errors.append(_error(field, value, "只能填写“是”或“否”。"))
    return default


def _date(value, field, errors):
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = _text(value)
    if not re.fullmatch(r"\d{4}-\d{2}-\d{2}", text):
        errors.append(_error(field, value, "日期必须使用 YYYY-MM-DD 格式。"))
        return None
    try:
        return date.fromisoformat(text)
    except ValueError:
        errors.append(_error(field, value, "日期不存在或格式无效。"))
        return None


def _read_uploaded(uploaded_file, limit):
    if getattr(uploaded_file, "size", 0) > limit:
        raise ValidationError(f"文件超过当前上限 {limit} 字节。")
    chunks = []
    total = 0
    temporary_path = None
    temporary_path_getter = getattr(uploaded_file, "temporary_file_path", None)
    if callable(temporary_path_getter):
        temporary_path = temporary_path_getter()

    from contextlib import nullcontext
    from django.conf import settings

    activity = (
        hold_temp_file_active(temporary_path, settings.IMPORT_TEMP_ROOT)
        if temporary_path
        else nullcontext()
    )
    with activity:
        for chunk in uploaded_file.chunks():
            total += len(chunk)
            if total > limit:
                raise ValidationError(f"文件超过当前上限 {limit} 字节。")
            chunks.append(chunk)
    if total <= 0:
        raise ValidationError("上传文件不能为空。")
    return b"".join(chunks)


def _validate_xlsx_container(data):
    if not data.startswith(b"PK\x03\x04"):
        raise ValidationError("文件内容不是有效的 XLSX ZIP 容器。")
    try:
        with zipfile.ZipFile(io.BytesIO(data)) as archive:
            members = archive.infolist()
            if len(members) > MAX_ARCHIVE_MEMBERS:
                raise ValidationError("XLSX 内部文件数量异常。")
            total_uncompressed = 0
            for item in members:
                if item.file_size < 0 or item.compress_size < 0:
                    raise ValidationError("XLSX ZIP 元数据无效。")
                if item.file_size > MAX_ARCHIVE_MEMBER_BYTES:
                    raise ValidationError("XLSX 单个内部文件解压后过大。")
                if item.file_size and item.compress_size == 0:
                    raise ValidationError("XLSX 内部文件压缩比异常。")
                if (
                    item.compress_size
                    and item.file_size / item.compress_size > MAX_COMPRESSION_RATIO
                ):
                    raise ValidationError("XLSX 内部文件压缩比异常。")
                total_uncompressed += item.file_size
            if total_uncompressed > MAX_UNCOMPRESSED_BYTES:
                raise ValidationError("XLSX 解压后大小超过安全上限。")
            # ZIP metadata is attacker-controlled.  Read every member through
            # a hard bounded loop as a second line of defence, instead of
            # trusting only the declared file_size/compress_size values.
            actual_total = 0
            for item in members:
                actual_member = 0
                with archive.open(item) as member_stream:
                    while chunk := member_stream.read(64 * 1024):
                        actual_member += len(chunk)
                        actual_total += len(chunk)
                        if actual_member > MAX_ARCHIVE_MEMBER_BYTES:
                            raise ValidationError(
                                "XLSX 单个内部文件解压后过大。"
                            )
                        if actual_total > MAX_UNCOMPRESSED_BYTES:
                            raise ValidationError(
                                "XLSX 解压后大小超过安全上限。"
                            )
            names = {item.filename.lower() for item in members}
            if "[content_types].xml" not in names or "xl/workbook.xml" not in names:
                raise ValidationError("文件缺少 XLSX 必需结构。")
            content_types = archive.read("[Content_Types].xml").lower()
            if b"macroenabled" in content_types or b"vbaproject" in content_types:
                raise ValidationError("文件声明了宏启用内容，已拒绝。")
            forbidden = sorted(
                name
                for name in names
                if any(part in name for part in FORBIDDEN_ARCHIVE_PARTS)
            )
            if forbidden:
                raise ValidationError("文件包含宏、外部链接或嵌入对象，已拒绝。")
            for item in members:
                if not item.filename.lower().endswith(".rels"):
                    continue
                try:
                    with archive.open(item) as relationship_stream:
                        relationships = DefusedElementTree.parse(
                            relationship_stream
                        ).getroot()
                except Exception as exc:
                    raise ValidationError("XLSX 关系 XML 无法安全解析。") from exc
                for relationship in relationships.iter():
                    if relationship.tag.rsplit("}", 1)[-1].casefold() != "relationship":
                        continue
                    attributes = [
                        (name.rsplit("}", 1)[-1].casefold(), str(value).strip())
                        for name, value in relationship.attrib.items()
                    ]
                    has_external_target = any(
                        name == "targetmode" and value.casefold() == "external"
                        for name, value in attributes
                    )
                    relationship_types = [
                        value.casefold() for name, value in attributes if name == "type"
                    ]
                    has_forbidden_type = any(
                        token in relationship_type
                        for relationship_type in relationship_types
                        for token in FORBIDDEN_RELATIONSHIP_TYPES
                    )
                    if has_external_target or has_forbidden_type:
                        raise ValidationError("文件包含外部关系或嵌入内容，已拒绝。")
            for item in members:
                name = item.filename.lower()
                if not name.startswith("xl/worksheets/") or not name.endswith(".xml"):
                    continue
                with archive.open(item) as worksheet_stream:
                    worksheet_errors = _validate_worksheet_xml(
                        worksheet_stream, item.filename
                    )
                if worksheet_errors:
                    return worksheet_errors
            return []
    except (zipfile.BadZipFile, zipfile.LargeZipFile) as exc:
        raise ValidationError("文件不是完整的 XLSX。") from exc


def _column_number(letters):
    number = 0
    for character in letters.upper():
        number = number * 26 + ord(character) - ord("A") + 1
    return number


def _validate_worksheet_xml(xml_source, worksheet_name):
    """Check actual cells rather than trusting a worksheet's dimension hint."""
    cell_count = 0
    if isinstance(xml_source, (bytes, bytearray)):
        xml_source = io.BytesIO(xml_source)
    try:
        for _event, element in DefusedElementTree.iterparse(
            xml_source, events=("end",)
        ):
            local_name = element.tag.rsplit("}", 1)[-1].lower()
            if local_name == "f":
                return [
                    _error(
                        "公式",
                        worksheet_name,
                        "所有工作表均禁止公式单元格。",
                    )
                ]
            if local_name == "c":
                cell_count += 1
                if cell_count > MAX_WORKSHEET_CELLS:
                    return [
                        _error(
                            "工作表",
                            worksheet_name,
                            "实际单元格数量超过安全上限："
                            f"最多 {MAX_WORKSHEET_CELLS} 个。",
                        )
                    ]
                reference = str(element.attrib.get("r", ""))
                match = CELL_REFERENCE_PATTERN.fullmatch(reference)
                if not match:
                    raise ValidationError(
                        f"工作表 {worksheet_name} 包含无效单元格坐标。"
                    )
                column_number = _column_number(match.group(1))
                row_number = int(match.group(2))
                if (
                    row_number > MAX_WORKSHEET_ROWS
                    or column_number > MAX_WORKSHEET_COLUMNS
                ):
                    return [
                        _error(
                            "工作表",
                            f"{worksheet_name}!{reference}",
                            "实际单元格超过安全上限："
                            f"最多 {MAX_WORKSHEET_ROWS} 行、{MAX_WORKSHEET_COLUMNS} 列。",
                        )
                    ]
            element.clear()
    except ValidationError:
        raise
    except Exception as exc:
        raise ValidationError("XLSX 工作表 XML 无法安全解析。") from exc
    return []


def _cell_is_formula(cell):
    return cell.data_type == "f" or (
        isinstance(cell.value, str) and cell.value.startswith("=")
    )


def _worksheet_limits_error(workbook):
    for worksheet in workbook.worksheets:
        if (
            worksheet.max_row > MAX_WORKSHEET_ROWS
            or worksheet.max_column > MAX_WORKSHEET_COLUMNS
        ):
            return _error(
                "工作表",
                worksheet.title,
                (
                    f"工作表维度超过安全上限：最多 {MAX_WORKSHEET_ROWS} 行、"
                    f"{MAX_WORKSHEET_COLUMNS} 列。"
                ),
            )
        for row in worksheet.iter_rows(
            min_row=1,
            max_row=worksheet.max_row,
            min_col=1,
            max_col=worksheet.max_column,
        ):
            for cell in row:
                if _cell_is_formula(cell):
                    return _error(
                        "公式",
                        f"{worksheet.title}!{cell.coordinate}",
                        "所有工作表均禁止公式单元格。",
                    )
    return None


def _xlsx_decimal_literals(data, worksheet_path):
    """Return numeric cell literals straight from XLSX XML, before float conversion."""

    literals = {}
    with zipfile.ZipFile(io.BytesIO(data)) as archive:
        with archive.open(worksheet_path) as source:
            for _event, element in DefusedElementTree.iterparse(
                source, events=("end",)
            ):
                if element.tag.rsplit("}", 1)[-1].lower() != "c":
                    continue
                reference = str(element.attrib.get("r", ""))
                cell_type = str(element.attrib.get("t", "") or "")
                if cell_type in {"", "n"}:
                    value_node = next(
                        (
                            child
                            for child in element
                            if child.tag.rsplit("}", 1)[-1].lower() == "v"
                        ),
                        None,
                    )
                    if value_node is not None and value_node.text is not None:
                        literals[reference] = value_node.text
                element.clear()
    return literals


def _load_rows(data, definition):
    try:
        workbook = load_workbook(
            io.BytesIO(data), read_only=True, data_only=False, keep_links=False
        )
    except Exception as exc:
        raise ValidationError("无法解析 XLSX，请重新下载标准模板。") from exc
    try:
        if definition.sheet_name not in workbook.sheetnames:
            return [], [
                _error(
                    "工作表",
                    ", ".join(workbook.sheetnames),
                    f"必须存在名为“{definition.sheet_name}”的工作表。",
                )
            ]
        expected_sheets = {definition.sheet_name, "填写说明"}
        if definition.has_example_sheet:
            expected_sheets.add("示例")
        missing_sheets = [name for name in expected_sheets if name not in workbook.sheetnames]
        if missing_sheets:
            return [], [
                _error(
                    "工作表",
                    ", ".join(workbook.sheetnames),
                    f"缺少标准工作表：{', '.join(sorted(missing_sheets))}。",
                )
            ]
        unknown_sheets = [name for name in workbook.sheetnames if name not in expected_sheets]
        if unknown_sheets:
            return [], [
                _error(
                    "工作表",
                    ", ".join(unknown_sheets),
                    "发现标准模板之外的工作表，已拒绝。",
                )
            ]
        safety_error = _worksheet_limits_error(workbook)
        if safety_error:
            return [], [safety_error]
        instructions = workbook["填写说明"]
        if _text(instructions["A2"].value) != "模板版本" or _text(
            instructions["B2"].value
        ) != definition.version:
            return [], [
                _error(
                    "模板版本",
                    instructions["B2"].value,
                    f"必须使用模板版本 {definition.version}。",
                )
            ]
        sheet = workbook[definition.sheet_name]
        numeric_literals = _xlsx_decimal_literals(data, sheet._worksheet_path)
        actual_headers = tuple(_text(cell.value) for cell in sheet[1])
        if actual_headers != definition.headers:
            unknown = [value for value in actual_headers if value not in definition.headers]
            missing = [value for value in definition.headers if value not in actual_headers]
            reason = "列名及顺序必须与标准模板完全一致。"
            if unknown:
                reason += f" 未知列：{', '.join(unknown)}。"
            if missing:
                reason += f" 缺少列：{', '.join(missing)}。"
            return [], [_error("表头", list(actual_headers), reason)]
        rows = []
        for row_number, cells in enumerate(
            sheet.iter_rows(min_row=2, max_col=len(definition.columns)), start=2
        ):
            if all(cell.value in (None, "") for cell in cells):
                continue
            formulas = [
                definition.columns[index].name
                for index, cell in enumerate(cells)
                if cell.data_type == "f"
                or (isinstance(cell.value, str) and cell.value.startswith("="))
            ]
            raw = {
                column.name: (
                    numeric_literals[cells[index].coordinate]
                    if isinstance(cells[index].value, float)
                    and cells[index].coordinate in numeric_literals
                    else _json_value(cells[index].value)
                )
                for index, column in enumerate(definition.columns)
            }
            values = {
                column.key: (
                    Decimal(numeric_literals[cells[index].coordinate])
                    if isinstance(cells[index].value, float)
                    and cells[index].coordinate in numeric_literals
                    else cells[index].value
                )
                for index, column in enumerate(definition.columns)
            }
            rows.append((row_number, values, raw, formulas))
            if len(rows) > MAX_IMPORT_ROWS:
                return [], [
                    _error(
                        "数据",
                        len(rows),
                        f"单批最多允许 {MAX_IMPORT_ROWS} 行有效数据。",
                    )
                ]
        return rows, []
    finally:
        workbook.close()


def _normalize_department_rows(company, loaded_rows, definition):
    from apps.masterdata.models import Department, Employee

    existing_codes = set(
        Department.objects.filter(company=company).values_list("normalized_code", flat=True)
    )
    employees = {
        employee.normalized_employee_no: employee
        for employee in Employee.objects.filter(company=company).select_related("department")
    }
    prepared = []
    seen = {}
    for row_number, values, raw, formulas in loaded_rows:
        errors = []
        if formulas:
            errors.extend(_error(field, raw.get(field), "禁止公式单元格。") for field in formulas)
        code = _text(values["code"])
        name = _text(values["name"])
        parent_code = _text(values["parent_code"])
        manager_no = _text(values["manager_employee_no"])
        normalized_code = normalize_identifier(code)
        normalized_parent = normalize_identifier(parent_code)
        normalized_manager = normalize_identifier(manager_no)
        if not code:
            errors.append(_error("部门编码", values["code"], "必填。"))
        if not name:
            errors.append(_error("部门名称", values["name"], "必填。"))
        if normalized_code in existing_codes:
            errors.append(_error("部门编码", code, "当前公司已存在该编码；本导入只新增。"))
        if normalized_code and normalized_code in seen:
            errors.append(_error("部门编码", code, f"与文件第 {seen[normalized_code]} 行重复。"))
        elif normalized_code:
            seen[normalized_code] = row_number
        if normalized_parent and normalized_parent == normalized_code:
            errors.append(_error("上级部门编码", parent_code, "不能将部门自身设为上级。"))
        manager = employees.get(normalized_manager) if normalized_manager else None
        if normalized_manager and manager is None:
            errors.append(_error("经理工号", manager_no, "当前公司不存在该员工编号。"))
        elif manager and (
            manager.employment_status != "active"
            or not manager.is_active
            or not manager.department.is_active
        ):
            errors.append(_error("经理工号", manager_no, "经理必须在职、启用且属于启用部门。"))
        normalized = {
            "code": code,
            "normalized_code": normalized_code,
            "name": name,
            "parent_code": parent_code,
            "normalized_parent_code": normalized_parent,
            "manager_employee_no": manager_no,
            "normalized_manager_employee_no": normalized_manager,
            "is_active": _boolean(values["is_active"], "是否启用", errors),
        }
        prepared.append({"row_number": row_number, "raw": raw, "normalized": normalized, "errors": errors})

    available = existing_codes | set(seen)
    parent_by_code = {item["normalized"]["normalized_code"]: item for item in prepared}
    for item in prepared:
        parent = item["normalized"]["normalized_parent_code"]
        if parent and parent not in available:
            item["errors"].append(_error("上级部门编码", item["normalized"]["parent_code"], "当前公司及本文件均不存在该部门编码。"))

    for item in prepared:
        start = item["normalized"]["normalized_code"]
        current = item["normalized"]["normalized_parent_code"]
        visited = {start}
        while current in parent_by_code:
            if current in visited:
                item["errors"].append(_error("上级部门编码", item["normalized"]["parent_code"], "文件内部门父级关系形成循环。"))
                break
            visited.add(current)
            current = parent_by_code[current]["normalized"]["normalized_parent_code"]
    return prepared


def _normalize_employee_rows(company, loaded_rows, definition):
    from apps.masterdata.models import Department, Employee

    existing_numbers = set(
        Employee.objects.filter(company=company).values_list("normalized_employee_no", flat=True)
    )
    departments = {
        department.normalized_code: department
        for department in Department.objects.filter(company=company)
    }
    prepared = []
    seen = {}
    for row_number, values, raw, formulas in loaded_rows:
        errors = []
        if formulas:
            errors.extend(_error(field, raw.get(field), "禁止公式单元格。") for field in formulas)
        employee_no = _text(values["employee_no"])
        name = _text(values["name"])
        department_code = _text(values["department_code"])
        status = _text(values["employment_status"])
        normalized_no = normalize_identifier(employee_no)
        normalized_department = normalize_identifier(department_code)
        if not employee_no:
            errors.append(_error("员工编号", values["employee_no"], "必填。"))
        if not name:
            errors.append(_error("姓名", values["name"], "必填。"))
        if not department_code:
            errors.append(_error("部门编码", values["department_code"], "必填。"))
        department = departments.get(normalized_department)
        if department is None:
            errors.append(_error("部门编码", department_code, "当前公司不存在该部门编码。"))
        elif not department.is_active:
            errors.append(_error("部门编码", department_code, "人员不能新绑定到已停用部门。"))
        if status not in {"active", "leaving", "resigned"}:
            errors.append(_error("任职状态", status, "只能为 active、leaving 或 resigned。"))
        if normalized_no in existing_numbers:
            errors.append(_error("员工编号", employee_no, "当前公司已存在该编号；本导入只新增。"))
        if normalized_no and normalized_no in seen:
            errors.append(_error("员工编号", employee_no, f"与文件第 {seen[normalized_no]} 行重复。"))
        elif normalized_no:
            seen[normalized_no] = row_number
        hire_date = _date(values["hire_date"], "入职日期", errors)
        termination_date = _date(values["termination_date"], "离职日期", errors)
        is_active = _boolean(values["is_active"], "是否启用", errors)
        if status == "resigned" and termination_date is None:
            errors.append(_error("离职日期", values["termination_date"], "resigned 人员必须填写离职日期。"))
        if status in {"active", "leaving"} and termination_date is not None:
            errors.append(_error("离职日期", values["termination_date"], "active/leaving 人员的离职日期必须为空。"))
        if status in {"leaving", "resigned"} and is_active:
            errors.append(_error("是否启用", values["is_active"], "leaving/resigned 人员必须停用。"))
        if hire_date and termination_date and termination_date < hire_date:
            errors.append(_error("离职日期", values["termination_date"], "离职日期不能早于入职日期。"))
        normalized = {
            "employee_no": employee_no,
            "normalized_employee_no": normalized_no,
            "name": name,
            "department_code": department_code,
            "normalized_department_code": normalized_department,
            "employment_status": status,
            "hire_date": hire_date.isoformat() if hire_date else None,
            "termination_date": termination_date.isoformat() if termination_date else None,
            "mobile": _text(values["mobile"]),
            "remark": _text(values["remark"]),
            "is_active": is_active,
        }
        prepared.append({"row_number": row_number, "raw": raw, "normalized": normalized, "errors": errors})
    return prepared


def _normalize_supply_item_rows(*, actor, company, loaded_rows, definition):
    from apps.supplies.models import SupplyCategory, SupplyItem, SupplyItemType, SupplyWarehouse
    from apps.supplies.permissions import can_manage_supply_item

    existing_codes = set(
        SupplyItem.objects.filter(company=company).values_list(
            "normalized_item_code", flat=True
        )
    )
    categories = {
        category.normalized_code: category
        for category in SupplyCategory.objects.filter(company=company, is_active=True)
    }
    warehouses = {
        warehouse.normalized_code: warehouse
        for warehouse in SupplyWarehouse.objects.filter(company=company, is_active=True)
    }
    prepared = []
    seen = {}
    for row_number, values, raw, formulas in loaded_rows:
        errors = []
        if formulas:
            errors.extend(
                _error(field, raw.get(field), "禁止公式单元格。")
                for field in formulas
            )
        item_code = _text(values["item_code"])
        name = _text(values["name"])
        category_code = _text(values["category_code"])
        item_type = _text(values["item_type"])
        unit = _text(values["unit"])
        warehouse_code = _text(values["default_warehouse_code"])
        normalized_code = normalize_identifier(item_code)
        normalized_category = normalize_identifier(category_code)
        normalized_warehouse = normalize_identifier(warehouse_code)
        category = categories.get(normalized_category)
        warehouse = warehouses.get(normalized_warehouse) if normalized_warehouse else None

        for label, value in (
            ("物品编码", item_code),
            ("物品名称", name),
            ("分类编码", category_code),
            ("管理模式", item_type),
            ("单位", unit),
        ):
            if not value:
                errors.append(_error(label, value, "必填。"))
        if normalized_code in existing_codes:
            errors.append(
                _error("物品编码", item_code, "当前公司已存在该编码；本导入只新增。")
            )
        if normalized_code and normalized_code in seen:
            errors.append(
                _error(
                    "物品编码",
                    item_code,
                    f"与文件第 {seen[normalized_code]} 行重复。",
                )
            )
        elif normalized_code:
            seen[normalized_code] = row_number
        if category is None:
            errors.append(
                _error("分类编码", category_code, "当前公司不存在该启用分类。")
            )
        if item_type not in SupplyItemType.values:
            errors.append(
                _error(
                    "管理模式",
                    item_type,
                    "只能为 consumable 或 durable_quantity。",
                )
            )
        elif not can_manage_supply_item(actor, item_type):
            errors.append(
                _error(
                    "管理模式",
                    item_type,
                    "当前角色无权导入该管理模式；equipment 仅可导入 durable_quantity。",
                )
            )
        if normalized_warehouse and warehouse is None:
            errors.append(
                _error(
                    "默认仓库编码",
                    warehouse_code,
                    "当前公司不存在该启用仓库。",
                )
            )
        minimum = _decimal_value(
            values["minimum_stock_quantity"],
            "最低库存",
            errors,
            places=4,
            minimum=Decimal("0"),
        )
        if minimum is None and values["minimum_stock_quantity"] in (None, ""):
            minimum = Decimal("0.0000")
        normalized = {
            "item_code": item_code,
            "normalized_item_code": normalized_code,
            "name": name,
            "category_id": str(category.pk) if category else None,
            "category_code": category_code,
            "normalized_category_code": normalized_category,
            "item_type": item_type,
            "unit": unit,
            "specification": _text(values["specification"]),
            "model": _text(values["model"]),
            "brand": _text(values["brand"]),
            "minimum_stock_quantity": str(minimum) if minimum is not None else None,
            "default_warehouse_id": str(warehouse.pk) if warehouse else None,
            "default_warehouse_code": warehouse_code,
            "normalized_default_warehouse_code": normalized_warehouse,
            "remark": _text(values["remark"]),
        }
        prepared.append(
            {
                "row_number": row_number,
                "raw": raw,
                "normalized": normalized,
                "errors": errors,
                "warnings": [],
            }
        )
    return prepared


def _inflate_supply_item_row(*, company, normalized, lock=False):
    from apps.supplies.models import SupplyCategory, SupplyWarehouse

    category_query = (
        SupplyCategory.objects.select_for_update()
        if lock
        else SupplyCategory.objects
    )
    warehouse_query = (
        SupplyWarehouse.objects.select_for_update()
        if lock
        else SupplyWarehouse.objects
    )
    category = category_query.filter(
        pk=normalized["category_id"],
        company=company,
        normalized_code=normalized["normalized_category_code"],
        is_active=True,
    ).first()
    if category is None:
        raise ValidationError("所选分类在确认前已停用、改码或不再属于当前公司。")
    warehouse = None
    if normalized["default_warehouse_id"]:
        warehouse = warehouse_query.filter(
            pk=normalized["default_warehouse_id"],
            company=company,
            normalized_code=normalized["normalized_default_warehouse_code"],
            is_active=True,
        ).first()
        if warehouse is None:
            raise ValidationError("默认仓库在确认前已停用、改码或不再属于当前公司。")
    return {
        "item_code": normalized["item_code"],
        "name": normalized["name"],
        "category": category,
        "item_type": normalized["item_type"],
        "unit": normalized["unit"],
        "specification": normalized["specification"],
        "model": normalized["model"],
        "brand": normalized["brand"],
        "minimum_stock_quantity": Decimal(normalized["minimum_stock_quantity"]),
        "default_warehouse": warehouse,
        "remark": normalized["remark"],
    }


def _normalize_opening_stock_rows(*, company, loaded_rows, definition):
    from apps.supplies.models import SupplyItem, SupplyWarehouse

    warehouses = {
        warehouse.normalized_code: warehouse
        for warehouse in SupplyWarehouse.objects.filter(company=company, is_active=True)
    }
    items = {
        item.normalized_item_code: item
        for item in SupplyItem.objects.filter(company=company, is_active=True)
    }
    prepared = []
    for row_number, values, raw, formulas in loaded_rows:
        errors = []
        if formulas:
            errors.extend(
                _error(field, raw.get(field), "禁止公式单元格。")
                for field in formulas
            )
        company_code = _text(values["company_code"])
        warehouse_code = _text(values["warehouse_code"])
        item_code = _text(values["item_code"])
        normalized_company = normalize_identifier(company_code)
        normalized_warehouse = normalize_identifier(warehouse_code)
        normalized_item = normalize_identifier(item_code)
        warehouse = warehouses.get(normalized_warehouse)
        item = items.get(normalized_item)
        for label, value in (
            ("公司编码", company_code),
            ("仓库编码", warehouse_code),
            ("物品编码", item_code),
        ):
            if not value:
                errors.append(_error(label, value, "必填。"))
        if normalized_company != company.normalized_code:
            errors.append(
                _error("公司编码", company_code, "必须精确匹配当前启用公司编码。")
            )
        if warehouse is None:
            errors.append(
                _error("仓库编码", warehouse_code, "当前公司不存在该启用仓库。")
            )
        if item is None:
            errors.append(
                _error("物品编码", item_code, "当前公司不存在该启用物品。")
            )
        quantity = _decimal_value(
            values["quantity"],
            "数量",
            errors,
            places=4,
            required=True,
            minimum=Decimal("0.0001"),
        )
        unit_cost = _decimal_value(
            values["unit_cost"],
            "单位成本",
            errors,
            places=6,
            required=True,
            minimum=Decimal("0"),
        )
        zero_cost_reason = _text(values["zero_cost_reason"])
        remark = _text(values["remark"])
        if unit_cost == Decimal("0") and not zero_cost_reason:
            errors.append(
                _error("0成本原因", values["zero_cost_reason"], "0 成本入库必须填写明确原因。")
            )
        normalized = {
            "company_code": company_code,
            "normalized_company_code": normalized_company,
            "warehouse_id": str(warehouse.pk) if warehouse else None,
            "warehouse_code": warehouse_code,
            "normalized_warehouse_code": normalized_warehouse,
            "item_id": str(item.pk) if item else None,
            "item_code": item_code,
            "normalized_item_code": normalized_item,
            "quantity": str(quantity) if quantity is not None else None,
            "unit_cost": str(unit_cost) if unit_cost is not None else None,
            "zero_cost_reason": zero_cost_reason,
            "remark": remark,
        }
        prepared.append(
            {
                "row_number": row_number,
                "raw": raw,
                "normalized": normalized,
                "errors": errors,
                "warnings": [],
            }
        )
    return prepared


def _inflate_opening_stock_row(*, company, normalized, lock=False):
    from apps.supplies.models import SupplyItem, SupplyWarehouse

    if normalized["normalized_company_code"] != company.normalized_code:
        raise ValidationError("公司编码在确认前已发生变化，请重新上传验证。")
    warehouse_query = (
        SupplyWarehouse.objects.select_for_update()
        if lock
        else SupplyWarehouse.objects
    )
    item_query = SupplyItem.objects.select_for_update() if lock else SupplyItem.objects
    warehouse = warehouse_query.filter(
        pk=normalized["warehouse_id"],
        company=company,
        normalized_code=normalized["normalized_warehouse_code"],
        is_active=True,
    ).first()
    if warehouse is None:
        raise ValidationError("仓库在确认前已停用、改码或不再属于当前公司。")
    item = item_query.filter(
        pk=normalized["item_id"],
        company=company,
        normalized_item_code=normalized["normalized_item_code"],
        is_active=True,
    ).first()
    if item is None:
        raise ValidationError("物品在确认前已停用、改码或不再属于当前公司。")
    reason = normalized["zero_cost_reason"]
    remark = normalized["remark"]
    line_parts = []
    if reason:
        line_parts.append(f"0成本原因：{reason}")
    if remark:
        line_parts.append(remark)
    return {
        "warehouse": warehouse,
        "item": item,
        "quantity": Decimal(normalized["quantity"]),
        "entered_unit_cost": Decimal(normalized["unit_cost"]),
        "line_remark": "；".join(line_parts),
    }


def _normalize_opening_custody_rows(*, company, loaded_rows, definition):
    from apps.masterdata.models import Department, Employee
    from apps.supplies.models import SupplyItem, SupplyItemType

    items = {
        item.normalized_item_code: item
        for item in SupplyItem.objects.filter(company=company, is_active=True)
    }
    departments = {
        department.normalized_code: department
        for department in Department.objects.filter(company=company, is_active=True)
    }
    employees = {
        employee.normalized_employee_no: employee
        for employee in Employee.objects.filter(company=company).select_related(
            "department"
        )
    }
    prepared = []
    for row_number, values, raw, formulas in loaded_rows:
        errors = []
        if formulas:
            errors.extend(
                _error(field, raw.get(field), "禁止公式单元格。")
                for field in formulas
            )
        item_code = _text(values["item_code"])
        department_code = _text(values["department_code"])
        employee_no = _text(values["employee_no"])
        normalized_item = normalize_identifier(item_code)
        normalized_department = normalize_identifier(department_code)
        normalized_employee = normalize_identifier(employee_no)
        item = items.get(normalized_item)
        department = departments.get(normalized_department)
        employee = employees.get(normalized_employee) if employee_no else None
        if not item_code:
            errors.append(_error("物品编码", values["item_code"], "必填。"))
        elif item is None:
            errors.append(_error("物品编码", item_code, "当前公司不存在该启用物品。"))
        elif item.item_type != SupplyItemType.DURABLE_QUANTITY:
            errors.append(_error("物品编码", item_code, "只允许数量型低值耐用品。"))
        if not department_code:
            errors.append(_error("责任部门编码", values["department_code"], "必填。"))
        elif department is None:
            errors.append(
                _error("责任部门编码", department_code, "当前公司不存在该启用部门。")
            )
        if employee_no:
            if employee is None:
                errors.append(
                    _error("责任员工编号", employee_no, "当前公司不存在该员工。")
                )
            else:
                if department is not None and employee.department_id != department.pk:
                    errors.append(
                        _error("责任员工编号", employee_no, "员工不属于责任部门。")
                    )
                if (
                    employee.employment_status != "active"
                    or not employee.is_active
                    or not employee.department.is_active
                ):
                    errors.append(
                        _error("责任员工编号", employee_no, "员工必须在职、启用且所属部门启用。")
                    )
        quantity = _decimal_value(
            values["quantity"],
            "数量",
            errors,
            places=4,
            required=True,
            minimum=Decimal("0.0001"),
        )
        unit_cost = _decimal_value(
            values["unit_cost"],
            "单位成本",
            errors,
            places=6,
            required=True,
            minimum=Decimal("0"),
        )
        started_on = _date(values["started_on"], "开始日期", errors)
        if values["started_on"] in (None, ""):
            errors.append(_error("开始日期", values["started_on"], "必填。"))
        normalized = {
            "item_id": str(item.pk) if item else None,
            "item_code": item_code,
            "normalized_item_code": normalized_item,
            "department_id": str(department.pk) if department else None,
            "department_code": department_code,
            "normalized_department_code": normalized_department,
            "employee_id": str(employee.pk) if employee else None,
            "employee_no": employee_no,
            "normalized_employee_no": normalized_employee,
            "quantity": str(quantity) if quantity is not None else None,
            "unit_cost": str(unit_cost) if unit_cost is not None else None,
            "started_on": started_on.isoformat() if started_on else None,
            "remark": _text(values["remark"]),
        }
        prepared.append(
            {
                "row_number": row_number,
                "raw": raw,
                "normalized": normalized,
                "errors": errors,
                "warnings": [],
            }
        )
    return prepared


def _inflate_opening_custody_row(*, company, normalized, lock=False):
    from apps.masterdata.models import Department, Employee
    from apps.supplies.models import SupplyItem, SupplyItemType

    item_query = SupplyItem.objects.select_for_update() if lock else SupplyItem.objects
    department_query = (
        Department.objects.select_for_update() if lock else Department.objects
    )
    employee_query = Employee.objects.select_related("department")
    if lock:
        employee_query = employee_query.select_for_update()
    item = item_query.filter(
        pk=normalized["item_id"],
        company=company,
        normalized_item_code=normalized["normalized_item_code"],
        item_type=SupplyItemType.DURABLE_QUANTITY,
        is_active=True,
    ).first()
    if item is None:
        raise ValidationError("物品在确认前已停用、改码或不再是数量型低值耐用品。")
    department = department_query.filter(
        pk=normalized["department_id"],
        company=company,
        normalized_code=normalized["normalized_department_code"],
        is_active=True,
    ).first()
    if department is None:
        raise ValidationError("责任部门在确认前已停用、改码或不再属于当前公司。")
    employee = None
    if normalized["employee_id"]:
        employee = employee_query.filter(
            pk=normalized["employee_id"],
            company=company,
            normalized_employee_no=normalized["normalized_employee_no"],
            department=department,
            employment_status="active",
            is_active=True,
            department__is_active=True,
        ).first()
        if employee is None:
            raise ValidationError("责任员工在确认前已离职、停用、改号或不再属于责任部门。")
    return {
        "item": item,
        "department": department,
        "employee": employee,
        "quantity": Decimal(normalized["quantity"]),
        "unit_cost": Decimal(normalized["unit_cost"]),
        "started_on": date.fromisoformat(normalized["started_on"]),
        "remark": normalized["remark"],
    }


def _asset_theoretical_summary(result, *, as_of_date=None):
    lines = result.get("lines", ()) if isinstance(result, dict) else result.lines
    cutoff_exclusive = as_of_date + timedelta(days=1) if as_of_date else None
    selected = [
        line
        for line in lines
        if cutoff_exclusive is None or line.period_end <= cutoff_exclusive
    ]
    if isinstance(result, dict):
        opening = result["opening_book_value"]
        salvage = result["salvage_value"]
        natural_end = result["natural_end_date"]
        start = result["start_date"]
        requires = result.get("requires_period_input")
    else:
        opening = result.opening_book_value
        salvage = result.salvage_value
        natural_end = result.natural_end_date
        start = result.start_date
        requires = None
    planned = sum((line.planned_amount for line in selected), Decimal("0.00"))
    closing = selected[-1].closing_book_value if selected else opening
    return _serialize_mapping(
        {
            "as_of_date": as_of_date,
            "start_date": start,
            "natural_end_date": natural_end,
            "salvage_value": salvage,
            "planned_accumulated_depreciation": planned,
            "theoretical_book_value": closing,
            "period_count": len(selected),
            "requires_period_input": requires,
        }
    )


def _inflate_asset_row(*, company, normalized, lock=False):
    from apps.assets.models import AssetCustomField
    from apps.masterdata.models import AssetCategory, Department, Employee, Location

    asset_data = normalized["asset_data"]
    query = lambda model: model.objects.select_for_update() if lock else model.objects
    category = query(AssetCategory).filter(
        pk=asset_data["category_id"], company=company, is_active=True
    ).first()
    department = query(Department).filter(
        pk=asset_data["department_id"], company=company, is_active=True
    ).first()
    employee = query(Employee).select_related("department").filter(
        pk=asset_data["responsible_employee_id"],
        company=company,
        employment_status="active",
        is_active=True,
        department__is_active=True,
    ).first()
    location = query(Location).filter(
        pk=asset_data["location_id"], company=company, is_active=True
    ).first()
    if not all((category, department, employee, location)):
        raise ValidationError("导入确认前的分类、部门、责任人或位置已失效。")
    if employee.department_id != department.pk:
        raise ValidationError("导入确认前责任人与资产部门已不一致。")
    if location.children.exists():
        raise ValidationError("导入确认前位置已不再是叶级。")
    data = {
        **asset_data,
        "category": category,
        "department": department,
        "responsible_employee": employee,
        "location": location,
        "acquisition_date": date.fromisoformat(asset_data["acquisition_date"])
        if asset_data["acquisition_date"]
        else None,
        "commissioning_date": date.fromisoformat(asset_data["commissioning_date"])
        if asset_data["commissioning_date"]
        else None,
    }
    for key in ("category_id", "department_id", "responsible_employee_id", "location_id"):
        data.pop(key, None)
    field_ids = set(normalized.get("custom_values", {}))
    fields = {
        str(value.pk): value
        for value in AssetCustomField.objects.filter(
            pk__in=field_ids, company=company, category=category, is_active=True
        )
    }
    if set(fields) != field_ids:
        raise ValidationError("导入确认前动态字段已停用或不再适用。")
    custom_values = {}
    for field_id, stored in normalized.get("custom_values", {}).items():
        field = fields[field_id]
        value = stored
        if field.field_type == "decimal":
            value = Decimal(stored)
        elif field.field_type == "date":
            value = date.fromisoformat(stored)
        custom_values[field_id] = value
    return data, custom_values


def _create_asset_finance_drafts(*, actor, asset, normalized, request=None):
    from apps.finance.models import (
        AssetDepreciationProfile,
        AssetFinance,
        DepreciationPolicy,
    )
    from apps.masterdata.models import FixedAssetCategory

    stored_finance = normalized.get("finance_data")
    if not stored_finance:
        return None, None
    if not can_manage_finance(actor):
        raise PermissionDenied("只有 finance 可确认带财务列的导入。")
    fixed_category = None
    if stored_finance.get("fixed_asset_category_id"):
        fixed_category = FixedAssetCategory.objects.filter(
            pk=stored_finance["fixed_asset_category_id"],
            company=asset.company,
            is_active=True,
        ).first()
        if fixed_category is None:
            raise ValidationError("导入确认前固定资产类别已失效。")
    stored_profile = normalized.get("profile_data")
    opening_impairment = Decimal("0.00")
    if stored_profile:
        opening_impairment = Decimal(
            stored_profile.get("opening_impairment") or "0.00"
        )
    finance = AssetFinance(
        company=asset.company,
        asset=asset,
        accounting_treatment=stored_finance["accounting_treatment"],
        accounting_treatment_reason=stored_finance["accounting_treatment_reason"],
        fixed_asset_category=fixed_category,
        original_cost=Decimal(stored_finance["original_cost"]),
        capitalization_date=date.fromisoformat(stored_finance["capitalization_date"])
        if stored_finance["capitalization_date"]
        else None,
        # This cache is an authoritative balance derived only from confirmed
        # adjustments.  The imported opening impairment remains in the draft
        # Profile/audit until Sprint 4 formalization posts its opening entry.
        impairment_balance_cache=Decimal("0.00"),
        finance_remark=stored_finance["finance_remark"],
    )
    finance.full_clean()
    finance.save()
    write_business_audit_log(
        company=asset.company,
        user=actor,
        action="asset_finance_import_draft_create",
        object_type="AssetFinance",
        object_id=finance.pk,
        old_data={},
        new_data={
            "asset_id": str(asset.pk),
            "accounting_treatment": finance.accounting_treatment,
            "fixed_asset_category_id": (
                str(finance.fixed_asset_category_id)
                if finance.fixed_asset_category_id
                else None
            ),
            "original_cost": str(finance.original_cost),
            "capitalization_date": finance.capitalization_date,
            "finance_confirmed": False,
        },
        **request_audit_context(request),
    )
    if not stored_profile:
        return finance, None
    profile_effective_from = date.fromisoformat(stored_profile["effective_from"])
    policy_business_date = timezone.localdate()
    policy = (
        DepreciationPolicy.objects.select_for_update()
        .filter(
            pk=stored_profile["depreciation_policy_id"],
            company=asset.company,
            status="active",
            effective_from__lte=policy_business_date,
        )
        .filter(Q(effective_to__isnull=True) | Q(effective_to__gte=policy_business_date))
        .first()
    )
    if policy is None:
        raise ValidationError("导入确认前折旧政策已失效或不再属于当前公司。")
    profile = AssetDepreciationProfile(
        company=asset.company,
        asset=asset,
        depreciation_policy=policy,
        version=1,
        method=stored_profile["method"],
        posting_period=stored_profile["posting_period"],
        start_rule=stored_profile["start_rule"],
        stop_rule=stored_profile["stop_rule"],
        start_date=date.fromisoformat(stored_profile["start_date"]),
        actual_continuation_date=date.fromisoformat(
            stored_profile["actual_continuation_date"]
        ),
        useful_life_months=stored_profile["useful_life_months"],
        salvage_mode=stored_profile["salvage_mode"],
        salvage_rate=Decimal(stored_profile["salvage_rate"])
        if stored_profile.get("salvage_rate") is not None
        else None,
        salvage_amount=Decimal(stored_profile["salvage_amount"])
        if stored_profile.get("salvage_amount") is not None
        else None,
        opening_book_value=Decimal(stored_profile["opening_book_value"]),
        opening_actual_accumulated_depreciation=Decimal(
            stored_profile["opening_actual_accumulated_depreciation"]
        ),
        expected_total_units=Decimal(stored_profile["expected_total_units"])
        if stored_profile.get("expected_total_units") is not None
        else None,
        work_unit=stored_profile.get("work_unit", ""),
        annual_posting_month=stored_profile.get("annual_posting_month"),
        effective_from=profile_effective_from,
        effective_to=date.fromisoformat(stored_profile["effective_to"])
        if stored_profile.get("effective_to")
        else None,
        status="draft",
        change_reason=stored_profile.get("change_reason", ""),
        created_by=actor,
    )
    profile.full_clean()
    profile.save()
    reference = normalized.get("theoretical_reference") or {}
    write_business_audit_log(
        company=asset.company,
        user=actor,
        action="depreciation_profile_import_draft_create",
        object_type="AssetDepreciationProfile",
        object_id=profile.pk,
        old_data={},
        new_data={
            "asset_id": str(asset.pk),
            "version": 1,
            "status": "draft",
            "opening_actual_accumulated_depreciation": str(
                profile.opening_actual_accumulated_depreciation
            ),
            "opening_impairment": str(opening_impairment),
            "opening_book_value": str(profile.opening_book_value),
            "actual_continuation_date": profile.actual_continuation_date,
            "theoretical_accumulated_depreciation": reference.get(
                "planned_accumulated_depreciation"
            ),
            "theoretical_difference": reference.get("difference"),
        },
        **request_audit_context(request),
    )
    return finance, profile


def _normalize_asset_rows(*, actor, company, loaded_rows, definition):
    from apps.assets.models import Asset, AssetCustomField
    from apps.finance.models import DepreciationPolicy
    from apps.finance.services import _profile_spec, resolve_depreciation_policy
    from apps.masterdata.models import (
        AssetCategory,
        Department,
        Employee,
        FixedAssetCategory,
        Location,
    )

    categories = {
        value.normalized_code: value
        for value in AssetCategory.objects.filter(company=company, is_active=True)
    }
    departments = {
        value.normalized_code: value
        for value in Department.objects.filter(company=company, is_active=True)
    }
    employees = {
        value.normalized_employee_no: value
        for value in Employee.objects.filter(
            company=company,
            employment_status="active",
            is_active=True,
            department__is_active=True,
        ).select_related("department")
    }
    locations = {
        value.normalized_code: value
        for value in Location.objects.filter(company=company, is_active=True)
    }
    fixed_categories = {
        value.normalized_code: value
        for value in FixedAssetCategory.objects.filter(company=company, is_active=True)
    }
    custom_fields = list(
        AssetCustomField.objects.filter(company=company, is_active=True).select_related(
            "category"
        )
    )
    custom_by_code = {value.normalized_code: value for value in custom_fields}
    finance_allowed = can_manage_finance(actor)
    warning_amount = get_system_setting(
        company=company, key="fixed_asset_warning_amount"
    )
    prepared = []
    for row_number, values, raw, formulas in loaded_rows:
        errors = []
        warnings = []
        if formulas:
            errors.extend(
                _error(field, raw.get(field), "禁止公式单元格。") for field in formulas
            )
        company_code = _text(values["company_code"])
        category_code = _text(values["category_code"])
        department_code = _text(values["department_code"])
        employee_no = _text(values["responsible_employee_no"])
        location_code = _text(values["location_code"])
        category = categories.get(normalize_identifier(category_code))
        department = departments.get(normalize_identifier(department_code))
        employee = employees.get(normalize_identifier(employee_no))
        location = locations.get(normalize_identifier(location_code))
        required_values = (
            ("资产名称", "asset_name"),
            ("实物分类编码", "category_code"),
            ("公司编码", "company_code"),
            ("部门编码", "department_code"),
            ("责任员工编号", "responsible_employee_no"),
            ("位置编码", "location_code"),
            ("单位", "unit"),
        )
        for label, key in required_values:
            if not _text(values[key]):
                errors.append(_error(label, values[key], "必填。"))
        if normalize_identifier(company_code) != company.normalized_code:
            errors.append(_error("公司编码", company_code, "必须精确匹配当前启用公司。"))
        if category is None:
            errors.append(_error("实物分类编码", category_code, "当前公司不存在该启用分类。"))
        if department is None:
            errors.append(_error("部门编码", department_code, "当前公司不存在该启用部门。"))
        elif not can_create_asset_draft(actor, company, department):
            errors.append(_error("部门编码", department_code, "超出调用人的资产草稿创建范围。"))
        if employee is None:
            errors.append(_error("责任员工编号", employee_no, "必须匹配当前公司在职、启用且所属部门启用的员工。"))
        elif department is not None and employee.department_id != department.pk:
            errors.append(_error("责任员工编号", employee_no, "责任人必须属于资产当前部门。"))
        if location is None:
            errors.append(_error("位置编码", location_code, "当前公司不存在该启用位置。"))
        elif location.children.exists():
            errors.append(_error("位置编码", location_code, "必须选择叶级具体位置。"))
        quantity = _integer_value(values["quantity"], "数量", errors, required=True)
        if quantity is not None and quantity != 1:
            errors.append(_error("数量", values["quantity"], "V1 每行只能为数量 1 的单件资产。"))
        acquisition_date = _date(values["acquisition_date"], "购置日期", errors)
        commissioning_date = _date(
            values["commissioning_date"], "达到可使用状态日期", errors
        )
        maintenance = _nullable_boolean(
            values["is_maintenance_required"], "是否需要保养", errors
        )
        if maintenance is None and category is not None:
            maintenance = category.is_maintenance_required_default
        attachment_note = _text(values["attachment_note"])
        if attachment_note and re.search(
            r"(?:https?://|file://|^[A-Za-z]:[\\/]|^[/\\])", attachment_note, re.I
        ):
            errors.append(_error("附件后续上传说明", attachment_note, "只允许后续受控上传说明，不接受本机路径或 URL。"))

        custom_values = {}
        for key, raw_value in values.items():
            if not key.startswith("custom:"):
                continue
            code = key.split(":", 1)[1]
            field = custom_by_code.get(code)
            if field is None:
                errors.append(_error(f"自定义:{code}", raw_value, "未知或已停用的动态字段。"))
                continue
            if category is None or field.category_id != category.pk:
                if raw_value not in (None, ""):
                    errors.append(_error(f"自定义:{field.code}", raw_value, "该字段不适用于本行实物分类。"))
                continue
            label = f"自定义:{field.code}"
            if raw_value in (None, ""):
                if field.required:
                    errors.append(_error(label, raw_value, "该分类的必填动态字段不得为空。"))
                continue
            parsed = raw_value
            if field.field_type == "decimal":
                parsed = _decimal_value(raw_value, label, errors, places=8)
            elif field.field_type == "date":
                parsed = _date(raw_value, label, errors)
            elif field.field_type == "boolean":
                parsed = _nullable_boolean(raw_value, label, errors)
            elif field.field_type in {"text", "select"}:
                parsed = _text(raw_value)
                if field.field_type == "select" and parsed not in field.options_json:
                    errors.append(_error(label, raw_value, "必须为已配置选项之一。"))
            if parsed is not None:
                custom_values[str(field.pk)] = _serialize_mapping(parsed)

        finance_supplied = any(values.get(key) not in (None, "") for key in ASSET_FINANCE_KEYS)
        finance_data = None
        profile_data = None
        theoretical_summary = None
        if finance_supplied and not finance_allowed:
            errors.append(_error("财务列", None, "只有 finance 可导入会计认定、期初值和折旧参数。"))
        elif finance_supplied:
            treatment = _text(values["accounting_treatment"])
            if treatment not in {"fixed_asset", "controlled_non_fixed"}:
                errors.append(_error("会计认定", treatment, "只能为 fixed_asset 或 controlled_non_fixed。"))
            cost = _decimal_value(values["original_cost"], "原值", errors, places=2, required=True, minimum=Decimal("0"))
            capitalization_date = _date(values["capitalization_date"], "资本化日期", errors)
            fixed_category_code = _text(values["fixed_asset_category_code"])
            fixed_category = fixed_categories.get(normalize_identifier(fixed_category_code))
            if treatment == "fixed_asset" and fixed_category is None:
                errors.append(_error("固定资产类别编码", fixed_category_code, "固定资产必须匹配当前公司启用会计类别。"))
            if treatment == "fixed_asset" and capitalization_date is None:
                errors.append(_error("资本化日期", values["capitalization_date"], "固定资产必填。"))
            if treatment == "controlled_non_fixed" and fixed_category_code:
                errors.append(_error("固定资产类别编码", fixed_category_code, "受控非固定资产不得填写固定资产类别。"))
            if treatment == "controlled_non_fixed":
                for key, label in (
                    ("depreciation_policy_key", "折旧政策编码"),
                    ("method", "折旧方法"),
                    ("posting_period", "计提周期"),
                    ("start_rule", "起算规则"),
                    ("specified_start", "指定起算日期"),
                    ("historical_start_reason", "历史起算原因"),
                    ("stop_rule", "停止规则"),
                    ("useful_life_months", "使用寿命月数"),
                    ("salvage_mode", "残值方式"),
                    ("salvage_rate", "残值率"),
                    ("salvage_amount", "残值金额"),
                    ("annual_posting_month", "年度计提月份"),
                    ("expected_total_units", "预计总工作量"),
                    ("work_unit", "工作量单位"),
                    ("opening_actual_accumulated_depreciation", "实际期初累计折旧"),
                    ("opening_impairment", "期初减值"),
                    ("opening_book_value", "实际期初账面净值"),
                    ("actual_continuation_date", "实际接续日"),
                    ("theoretical_as_of_date", "理论测算截止日"),
                ):
                    if values[key] not in (None, ""):
                        errors.append(
                            _error(
                                label,
                                values[key],
                                "受控非固定资产不得填写折旧或期初字段。",
                            )
                        )
            finance_data = {
                "accounting_treatment": treatment,
                "accounting_treatment_reason": _text(values["accounting_treatment_reason"]),
                "fixed_asset_category_id": str(fixed_category.pk) if fixed_category else None,
                "original_cost": str(cost) if cost is not None else None,
                "capitalization_date": capitalization_date.isoformat() if capitalization_date else None,
                "finance_remark": _text(values["finance_remark"]),
            }
            if cost is not None and cost >= warning_amount:
                warnings.append(_error("原值", str(cost), f"达到固定资产提示阈值 {warning_amount}，仅提醒，不自动改变会计认定。"))
            if treatment == "fixed_asset" and category and commissioning_date and cost is not None and fixed_category:
                policy_key = _text(values["depreciation_policy_key"])
                requested = None
                if policy_key:
                    policy_business_date = timezone.localdate()
                    policies = list(
                        DepreciationPolicy.objects.filter(
                            company=company,
                            policy_key=policy_key,
                            status="active",
                            effective_from__lte=policy_business_date,
                        ).filter(
                            Q(effective_to__isnull=True)
                            | Q(effective_to__gte=policy_business_date)
                        )
                    )
                    if len(policies) != 1:
                        errors.append(_error("折旧政策编码", policy_key, "必须精确解析一个当前生效政策版本。"))
                    else:
                        requested = policies[0]
                temp_asset = Asset(company=company, category=category, commissioning_date=commissioning_date)
                profile_values = {
                    "method": _text(values["method"]) or None,
                    "posting_period": _text(values["posting_period"]) or None,
                    "start_rule": _text(values["start_rule"]) or None,
                    "specified_start": _date(values["specified_start"], "指定起算日期", errors),
                    "stop_rule": _text(values["stop_rule"]) or None,
                    "useful_life_months": _integer_value(values["useful_life_months"], "使用寿命月数", errors, minimum=1),
                    "salvage_mode": _text(values["salvage_mode"]) or None,
                    "salvage_rate": _decimal_value(values["salvage_rate"], "残值率", errors, places=8, minimum=Decimal("0")),
                    "salvage_amount": _decimal_value(values["salvage_amount"], "残值金额", errors, places=2, minimum=Decimal("0")),
                    "annual_posting_month": _integer_value(values["annual_posting_month"], "年度计提月份", errors, minimum=1, maximum=12),
                    "expected_total_units": _decimal_value(values["expected_total_units"], "预计总工作量", errors, places=6, minimum=Decimal("0")),
                    "work_unit": _text(values["work_unit"]),
                    "opening_actual_accumulated_depreciation": _decimal_value(values["opening_actual_accumulated_depreciation"], "实际期初累计折旧", errors, places=2, minimum=Decimal("0")) or Decimal("0.00"),
                    "opening_impairment": _decimal_value(values["opening_impairment"], "期初减值", errors, places=2, minimum=Decimal("0")) or Decimal("0.00"),
                    "opening_book_value": _decimal_value(values["opening_book_value"], "实际期初账面净值", errors, places=2, minimum=Decimal("0")),
                    "actual_continuation_date": _date(
                        values["actual_continuation_date"], "实际接续日", errors
                    ),
                }
                profile_values = {key: value for key, value in profile_values.items() if value is not None}
                historical_reason = _text(values["historical_start_reason"])
                if historical_reason:
                    profile_values["allow_historical_start"] = True
                    profile_values["change_reason"] = historical_reason
                if (
                    historical_reason
                    or profile_values["opening_actual_accumulated_depreciation"]
                    != Decimal("0.00")
                    or profile_values["opening_impairment"] != Decimal("0.00")
                ) and profile_values.get("actual_continuation_date") is None:
                    errors.append(
                        _error(
                            "实际接续日",
                            values["actual_continuation_date"],
                            "旧资产初始化必须填写独立的实际接续日。",
                        )
                    )
                try:
                    policy = resolve_depreciation_policy(asset=temp_asset, requested_policy=requested)
                    _spec, result, resolved = _profile_spec(asset=temp_asset, finance_data={"original_cost": cost, "fixed_asset_category": fixed_category}, profile_data=profile_values, policy=policy)
                    resolved["depreciation_policy_id"] = str(policy.pk)
                    resolved["opening_impairment"] = profile_values.get("opening_impairment", Decimal("0.00"))
                    profile_data = _serialize_mapping(resolved)
                    theoretical_date = _date(values["theoretical_as_of_date"], "理论测算截止日", errors)
                    theoretical_inputs = {
                        **profile_values,
                        "opening_actual_accumulated_depreciation": Decimal("0.00"),
                        "opening_impairment": Decimal("0.00"),
                        "opening_book_value": cost,
                    }
                    theoretical_inputs["actual_continuation_date"] = resolved[
                        "start_date"
                    ]
                    _theoretical_spec, theoretical_result, _theoretical_resolved = _profile_spec(
                        asset=temp_asset,
                        finance_data={
                            "original_cost": cost,
                            "fixed_asset_category": fixed_category,
                        },
                        profile_data=theoretical_inputs,
                        policy=policy,
                    )
                    theoretical_summary = _asset_theoretical_summary(
                        theoretical_result, as_of_date=theoretical_date
                    )
                    theoretical_summary["actual_opening_book_value"] = profile_data["opening_book_value"]
                    theoretical_summary["difference"] = str(Decimal(profile_data["opening_book_value"]) - Decimal(theoretical_summary["theoretical_book_value"]))
                except (ValidationError, ValueError) as exc:
                    errors.extend(_error("折旧参数", None, message) for message in _validation_messages(exc))

        normalized = {
            "asset_data": {
                "asset_name": _text(values["asset_name"]),
                "category_id": str(category.pk) if category else None,
                "brand": _text(values["brand"]),
                "model": _text(values["model"]),
                "manufacturer": _text(values["manufacturer"]),
                "serial_number": _text(values["serial_number"]),
                "factory_number": _text(values["factory_number"]),
                "historical_code": _text(values["historical_code"]),
                "unit": _text(values["unit"]),
                "description": attachment_note,
                "department_id": str(department.pk) if department else None,
                "responsible_employee_id": str(employee.pk) if employee else None,
                "location_id": str(location.pk) if location else None,
                "acquisition_date": acquisition_date.isoformat() if acquisition_date else None,
                "commissioning_date": commissioning_date.isoformat() if commissioning_date else None,
                "is_maintenance_required": bool(maintenance),
                "notes": _text(values["notes"]),
            },
            "custom_values": custom_values,
            "finance_data": finance_data,
            "profile_data": profile_data,
            "theoretical_reference": theoretical_summary,
        }
        prepared.append({"row_number": row_number, "raw": raw, "normalized": normalized, "errors": errors, "warnings": warnings})

    # Potential identity duplicates are warnings, never silent deduplication.
    identity_fields = (("serial_number", "序列号"), ("factory_number", "出厂编号"), ("historical_code", "历史参考编号"))
    for key, label in identity_fields:
        seen = {}
        db_values = set(Asset.objects.filter(company=company).exclude(**{key: ""}).values_list(key, flat=True))
        for item in prepared:
            value = item["normalized"]["asset_data"][key]
            if not value:
                continue
            if value in db_values:
                item["warnings"].append(_error(label, value, "当前公司已有相同值，请人工核查潜在重复。"))
            if value in seen:
                item["warnings"].append(_error(label, value, f"与文件第 {seen[value]} 行相同，请人工核查。"))
            else:
                seen[value] = item["row_number"]
    return prepared


def _audit_batch(*, batch, actor, action, new_data, old_data=None, request=None):
    write_business_audit_log(
        company=batch.company,
        user=actor,
        action=action,
        object_type="ImportBatch",
        object_id=batch.pk,
        old_data=old_data or {},
        new_data=new_data,
        **request_audit_context(request),
    )


def _validation_messages(exc):
    if hasattr(exc, "message_dict"):
        messages = []
        for field, field_messages in exc.message_dict.items():
            messages.extend(f"{field}: {message}" for message in field_messages)
        return messages
    return list(getattr(exc, "messages", (str(exc),)))


def _preflight_business_rows(*, actor, company, import_type, prepared):
    """Run the same model/service validation without leaving business rows behind."""

    from apps.masterdata.models import Department, Employee

    if any(item["errors"] for item in prepared):
        return prepared
    with transaction.atomic():
        if import_type == "asset_initialization":
            for item in prepared:
                try:
                    data, custom_values = _inflate_asset_row(
                        company=company,
                        normalized=item["normalized"],
                    )
                    asset = create_asset_draft(
                        actor=actor,
                        company=company,
                        data=data,
                        custom_values=custom_values,
                        initialization_source="excel_import",
                    )
                    _create_asset_finance_drafts(
                        actor=actor,
                        asset=asset,
                        normalized=item["normalized"],
                    )
                except (ValidationError, PermissionDenied) as exc:
                    item["errors"].extend(
                        _error("数据", None, message)
                        for message in _validation_messages(exc)
                    )
        elif import_type == "item_master":
            from apps.supplies.services import create_supply_item

            for item in prepared:
                try:
                    create_supply_item(
                        actor=actor,
                        company=company,
                        data=_inflate_supply_item_row(
                            company=company,
                            normalized=item["normalized"],
                        ),
                    )
                except (ValidationError, PermissionDenied) as exc:
                    item["errors"].extend(
                        _error("数据", None, message)
                        for message in _validation_messages(exc)
                    )
        elif import_type == "opening_stock":
            from apps.supplies.services import create_supply_document

            groups = {}
            for item in prepared:
                try:
                    row_data = _inflate_opening_stock_row(
                        company=company,
                        normalized=item["normalized"],
                    )
                except (ValidationError, PermissionDenied) as exc:
                    item["errors"].extend(
                        _error("数据", None, message)
                        for message in _validation_messages(exc)
                    )
                    continue
                groups.setdefault(row_data["warehouse"].pk, []).append(
                    (item, row_data)
                )
            for grouped in groups.values():
                try:
                    create_supply_document(
                        actor=actor,
                        company=company,
                        document_type="opening",
                        data={
                            "business_date": timezone.localdate(),
                            "target_warehouse": grouped[0][1]["warehouse"],
                            "idempotency_key": f"opening-preflight-{uuid.uuid4()}",
                            "remark": "期初库存导入预检",
                        },
                        lines=[
                            {
                                "item": row_data["item"],
                                "quantity": row_data["quantity"],
                                "entered_unit_cost": row_data["entered_unit_cost"],
                                "line_remark": row_data["line_remark"],
                            }
                            for _, row_data in grouped
                        ],
                    )
                except (ValidationError, PermissionDenied) as exc:
                    for item, _ in grouped:
                        item["errors"].extend(
                            _error("数据", None, message)
                            for message in _validation_messages(exc)
                        )
        elif import_type == "opening_custody":
            for item in prepared:
                try:
                    _inflate_opening_custody_row(
                        company=company,
                        normalized=item["normalized"],
                    )
                except (ValidationError, PermissionDenied) as exc:
                    item["errors"].extend(
                        _error("数据", None, message)
                        for message in _validation_messages(exc)
                    )
        elif import_type == "department":
            existing = {
                value.normalized_code: value
                for value in Department.objects.filter(company=company)
            }
            employees = {
                value.normalized_employee_no: value
                for value in Employee.objects.filter(company=company).select_related(
                    "department"
                )
            }
            created = {}
            pending = list(prepared)
            while pending:
                progressed = False
                for item in pending[:]:
                    normalized = item["normalized"]
                    parent_code = normalized["normalized_parent_code"]
                    parent = existing.get(parent_code) or created.get(parent_code)
                    if parent_code and parent is None:
                        continue
                    try:
                        department = create_department(
                            actor=actor,
                            company=company,
                            data={
                                "code": normalized["code"],
                                "name": normalized["name"],
                                "parent": parent,
                                "manager_employee": employees.get(
                                    normalized["normalized_manager_employee_no"]
                                ),
                                "is_active": normalized["is_active"],
                            },
                        )
                    except ValidationError as exc:
                        item["errors"].extend(
                            _error("数据", None, message)
                            for message in _validation_messages(exc)
                        )
                        department = None
                    if department is not None:
                        created[normalized["normalized_code"]] = department
                    pending.remove(item)
                    progressed = True
                if not progressed:
                    break
        elif import_type == "employee":
            departments = {
                value.normalized_code: value
                for value in Department.objects.filter(company=company, is_active=True)
            }
            for item in prepared:
                normalized = item["normalized"]
                try:
                    create_employee(
                        actor=actor,
                        company=company,
                        data={
                            "employee_no": normalized["employee_no"],
                            "name": normalized["name"],
                            "department": departments.get(
                                normalized["normalized_department_code"]
                            ),
                            "employment_status": normalized["employment_status"],
                            "hire_date": date.fromisoformat(normalized["hire_date"])
                            if normalized["hire_date"]
                            else None,
                            "termination_date": date.fromisoformat(
                                normalized["termination_date"]
                            )
                            if normalized["termination_date"]
                            else None,
                            "mobile": normalized["mobile"],
                            "remark": normalized["remark"],
                            "is_active": normalized["is_active"],
                        },
                    )
                except ValidationError as exc:
                    item["errors"].extend(
                        _error("数据", None, message)
                        for message in _validation_messages(exc)
                    )
        else:
            raise ValidationError("不支持的导入类型。")
        transaction.set_rollback(True)
    return prepared


def upload_and_validate_import(
    *, actor, company, import_type, uploaded_file, idempotency_key, request=None
):
    from apps.masterdata.models import Attachment, ImportBatch, ImportRow

    require_import_permission(actor, import_type, company=company)
    _require_current_import_company(company)
    definition = get_template_definition(import_type, company=company)
    allowed = set(get_system_setting(company=company, key="attachment_allowed_extensions"))
    if "xlsx" not in allowed:
        raise ValidationError("当前公司附件白名单未允许 xlsx。")
    original_filename, extension = _validate_filename(uploaded_file.name)
    if extension != "xlsx":
        raise ValidationError("只允许上传无宏的 .xlsx 文件。")
    content_type = str(getattr(uploaded_file, "content_type", "") or "").lower()
    if content_type != XLSX_MIME:
        raise ValidationError("上传文件的 MIME 类型不是标准 XLSX。")
    limit = get_system_setting(company=company, key="attachment_max_size_bytes")
    data = _read_uploaded(uploaded_file, limit)
    container_errors = _validate_xlsx_container(data)
    digest = hashlib.sha256(data).hexdigest()
    request_hash = hashlib.sha256(
        f"{import_type}:{definition.version}:{digest}".encode()
    ).hexdigest()

    with transaction.atomic():
        _lock_import_namespace(company.pk)
        existing_key = (
            ImportBatch.objects.select_for_update()
            .filter(company=company, idempotency_key=idempotency_key)
            .first()
        )
        if existing_key:
            if existing_key.request_hash != request_hash:
                raise ValidationError("同一幂等标识已用于不同文件。")
            return existing_key
        duplicate = (
            ImportBatch.objects.select_for_update()
            .filter(
                company=company,
                import_type=import_type,
                template_version=definition.version,
                file_sha256=digest,
            )
            .order_by("-uploaded_at")
            .first()
        )
        if duplicate:
            raise ValidationError(
                f"该文件已上传为批次 {duplicate.pk}，请打开原批次，不要重复上传。"
            )

    if container_errors:
        loaded_rows, workbook_errors = [], container_errors
    else:
        loaded_rows, workbook_errors = _load_rows(data, definition)
    if workbook_errors:
        prepared = [{"row_number": 1, "raw": {}, "normalized": {}, "errors": workbook_errors}]
    elif import_type == "department":
        prepared = _normalize_department_rows(company, loaded_rows, definition)
    elif import_type == "employee":
        prepared = _normalize_employee_rows(company, loaded_rows, definition)
    elif import_type == "item_master":
        prepared = _normalize_supply_item_rows(
            actor=actor,
            company=company,
            loaded_rows=loaded_rows,
            definition=definition,
        )
    elif import_type == "opening_stock":
        prepared = _normalize_opening_stock_rows(
            company=company,
            loaded_rows=loaded_rows,
            definition=definition,
        )
    elif import_type == "opening_custody":
        prepared = _normalize_opening_custody_rows(
            company=company,
            loaded_rows=loaded_rows,
            definition=definition,
        )
    elif import_type == "asset_initialization":
        prepared = _normalize_asset_rows(
            actor=actor,
            company=company,
            loaded_rows=loaded_rows,
            definition=definition,
        )
    else:
        raise ValidationError("不支持的导入类型。")
    prepared = _preflight_business_rows(
        actor=actor,
        company=company,
        import_type=import_type,
        prepared=prepared,
    )
    if not prepared:
        prepared = [{"row_number": 2, "raw": {}, "normalized": {}, "errors": [_error("数据", None, "模板中没有可导入的数据行。")]}]

    # The object is written first under a private, random key.  It has no
    # public URL and is not downloadable until the database transaction below
    # atomically publishes Attachment.is_available=True.  A process failure
    # before the database commit therefore leaves, at worst, an unreferenced
    # private object for the idempotent cleanup command.
    storage_key = f"private/imports/{company.pk}/{uuid.uuid4().hex}.xlsx"
    saved_key = default_storage.save(storage_key, ContentFile(data))
    batch = None
    created_batch = False
    try:
        with transaction.atomic():
            _lock_import_namespace(company.pk)
            existing_key = (
                ImportBatch.objects.select_for_update()
                .filter(company=company, idempotency_key=idempotency_key)
                .first()
            )
            if existing_key:
                if existing_key.request_hash != request_hash:
                    raise ValidationError("同一幂等标识已用于不同文件。")
                batch = existing_key
            if batch is not None:
                # Do not create a second evidence attachment for an idempotent
                # retry.  The just-written private object is removed after the
                # transaction has released its locks.
                pass
            else:
                duplicate = (
                    ImportBatch.objects.select_for_update()
                    .filter(
                        company=company,
                        import_type=import_type,
                        template_version=definition.version,
                        file_sha256=digest,
                    )
                    .order_by("-uploaded_at")
                    .first()
                )
                if duplicate:
                    raise ValidationError(
                        f"该文件已上传为批次 {duplicate.pk}，请打开原批次，不要重复上传。"
                    )
                attachment = Attachment(
                    company=company,
                    storage_key=saved_key,
                    original_filename=original_filename[:255],
                    safe_filename=(get_valid_filename(original_filename) or "import.xlsx")[:255],
                    file_size=len(data),
                    mime_type=XLSX_MIME,
                    sha256=digest,
                    uploaded_by=actor,
                    malware_scan_status="pending",
                    is_available=False,
                )
                attachment.full_clean()
                attachment.save()
                errors_count = sum(bool(item["errors"]) for item in prepared)
                warning_count = sum(bool(item.get("warnings")) for item in prepared)
                batch = ImportBatch(
                    company=company,
                    import_type=import_type,
                    template_version=definition.version,
                    file_attachment=attachment,
                    file_sha256=digest,
                    status="invalid" if errors_count else "validated",
                    total_rows=len(prepared),
                    valid_rows=len(prepared) - errors_count,
                    error_rows=errors_count,
                    warning_rows=warning_count,
                    request_hash=request_hash,
                    idempotency_key=idempotency_key,
                    uploaded_by=actor,
                    validated_at=timezone.now(),
                )
                batch.full_clean()
                batch.save()
                created_batch = True
                rows = [
                    ImportRow(
                        batch=batch,
                        row_number=item["row_number"],
                        raw_data_json=item["raw"],
                        normalized_data_json=item["normalized"],
                        validation_status="invalid" if item["errors"] else "valid",
                        errors_json=item["errors"],
                        warnings_json=item.get("warnings", []),
                    )
                    for item in prepared
                ]
                for row in rows:
                    row.full_clean()
                ImportRow.objects.bulk_create(rows)
                _audit_batch(
                    batch=batch,
                    actor=actor,
                    action="import_upload",
                    new_data={
                        "import_type": import_type,
                        "template_version": definition.version,
                        "file_sha256": digest,
                        "status": "uploaded",
                    },
                    request=request,
                )
                _audit_batch(
                    batch=batch,
                    actor=actor,
                    action="import_validate",
                    new_data={
                        "import_type": import_type,
                        "template_version": definition.version,
                        "file_sha256": digest,
                        "status": batch.status,
                        "total_rows": batch.total_rows,
                        "valid_rows": batch.valid_rows,
                        "error_rows": batch.error_rows,
                        "warning_rows": batch.warning_rows,
                    },
                    request=request,
                )
                if errors_count:
                    _audit_batch(
                        batch=batch,
                        actor=actor,
                        action="import_failure",
                        new_data={
                            "import_type": import_type,
                            "template_version": definition.version,
                            "file_sha256": digest,
                            "status": batch.status,
                            "stage": "validation",
                            "error_rows": batch.error_rows,
                        },
                        request=request,
                    )
                # This update is the publication point.  Other transactions
                # can only observe it after the validation rows and audit log
                # have committed successfully.
                attachment.malware_scan_status = "policy_limited"
                attachment.is_available = True
                attachment.full_clean()
                attachment.save(
                    update_fields=["malware_scan_status", "is_available"]
                )
        if not created_batch:
            default_storage.delete(saved_key)
        return batch
    except Exception:
        if saved_key:
            try:
                default_storage.delete(saved_key)
            except Exception:
                # Never mask the database/validation failure.  The private
                # unreferenced object is handled by cleanup_import_staging.
                pass
        raise


@transaction.atomic
def cancel_import_batch(*, actor, batch, reason, request=None):
    from apps.masterdata.models import ImportBatch

    cleaned_reason = str(reason or "").strip()
    if not cleaned_reason:
        raise ValidationError({"reason": "取消导入批次必须填写原因。"})
    _lock_import_namespace(batch.company_id)
    batch = (
        ImportBatch.objects.select_for_update()
        .select_related("company")
        .get(pk=batch.pk)
    )
    _require_current_import_company(batch.company)
    require_import_permission(actor, batch.import_type, company=batch.company)
    if batch.status == "cancelled":
        return batch
    if batch.status == "confirmed":
        if batch.import_type == "opening_custody":
            raise ValidationError("耐用品期初保管导入确认后不得破坏性回滚或取消。")
        raise ValidationError("已确认导入批次不得取消。")
    if batch.status not in {"uploaded", "validated", "invalid", "failed"}:
        raise ValidationError("当前导入批次状态不能取消。")
    old_status = batch.status
    batch.status = "cancelled"
    batch.full_clean()
    batch.save(update_fields=["status"])
    _audit_batch(
        batch=batch,
        actor=actor,
        action="import_cancel",
        old_data={"status": old_status},
        new_data={
            "status": "cancelled",
            "import_type": batch.import_type,
            "reason": cleaned_reason,
        },
        request=request,
    )
    return batch


def confirm_import_batch(*, actor, batch, request=None):
    """Confirm atomically and persist a separate safe failure audit on rollback."""

    try:
        return _confirm_import_batch_atomic(
            actor=actor, batch=batch, request=request
        )
    except Exception as exc:
        from apps.masterdata.models import ImportBatch

        persisted = (
            ImportBatch.objects.select_related("company")
            .filter(pk=getattr(batch, "pk", None))
            .first()
        )
        if persisted is not None:
            try:
                _audit_batch(
                    batch=persisted,
                    actor=actor,
                    action="import_confirm_failed",
                    old_data={"status": persisted.status},
                    new_data={
                        "status": persisted.status,
                        "import_type": persisted.import_type,
                        "file_sha256": persisted.file_sha256,
                        "stage": "confirmation",
                        "error_class": type(exc).__name__,
                    },
                    request=request,
                )
            except Exception:
                # The original confirmation exception is the authoritative
                # caller result.  Never replace it with a secondary audit
                # transport failure.
                pass
        raise


@transaction.atomic
def _confirm_import_batch_atomic(*, actor, batch, request=None):
    from apps.masterdata.models import Department, Employee, ImportBatch

    # Share the upload/cleanup namespace lock.  Cleanup can therefore prove
    # that no confirmation/idempotency request is processing the company while
    # it re-checks and removes an expired staging batch.
    _lock_import_namespace(batch.company_id)
    batch = (
        ImportBatch.objects.select_for_update()
        .select_related("company", "file_attachment")
        .get(pk=batch.pk)
    )
    _require_current_import_company(batch.company)
    require_import_permission(actor, batch.import_type, company=batch.company)
    if batch.status == "confirmed":
        return batch
    definition = get_template_definition(batch.import_type, company=batch.company)
    if batch.template_version != definition.version:
        raise ValidationError("批次模板版本已不再受支持，请重新下载模板并上传。")
    if batch.status != "validated" or batch.error_rows != 0:
        raise ValidationError("只能确认无错误的已验证批次。")
    if batch.file_sha256 != batch.file_attachment.sha256:
        raise ValidationError("原文件摘要与批次不一致，已阻止确认。")
    if not batch.file_attachment.is_available:
        raise ValidationError("原文件已不可用，已阻止确认。")
    if batch.file_attachment.malware_scan_status not in {
        "policy_limited",
        "clean",
    }:
        raise ValidationError("原文件尚未通过安全策略校验，已阻止确认。")
    rows = list(batch.rows.select_for_update().order_by("row_number"))
    if not rows or any(row.validation_status != "valid" for row in rows):
        raise ValidationError("批次行状态与已验证状态不一致。")

    created = {}
    if batch.import_type == "asset_initialization":
        for row in rows:
            item = row.normalized_data_json
            data, custom_values = _inflate_asset_row(
                company=batch.company, normalized=item, lock=True
            )
            require_import_permission(
                actor,
                batch.import_type,
                company=batch.company,
                department=data["department"],
            )
            asset = create_asset_draft(
                actor=actor,
                company=batch.company,
                data=data,
                custom_values=custom_values,
                initialization_source="excel_import",
                request=request,
            )
            _create_asset_finance_drafts(
                actor=actor,
                asset=asset,
                normalized=item,
                request=request,
            )
            row.created_object_type = "Asset"
            row.created_object_id = str(asset.pk)
    elif batch.import_type == "item_master":
        from apps.supplies.models import SupplyItem
        from apps.supplies.services import create_supply_item

        existing_codes = set(
            SupplyItem.objects.select_for_update()
            .filter(company=batch.company)
            .values_list("normalized_item_code", flat=True)
        )
        if any(
            row.normalized_data_json["normalized_item_code"] in existing_codes
            for row in rows
        ):
            raise ValidationError("确认前检测到物品编码已存在，请重新上传验证。")
        for row in rows:
            item = create_supply_item(
                actor=actor,
                company=batch.company,
                data=_inflate_supply_item_row(
                    company=batch.company,
                    normalized=row.normalized_data_json,
                    lock=True,
                ),
                request=request,
            )
            row.created_object_type = "SupplyItem"
            row.created_object_id = str(item.pk)
    elif batch.import_type == "opening_stock":
        from apps.supplies.services import create_supply_document

        grouped_rows = {}
        for row in rows:
            row_data = _inflate_opening_stock_row(
                company=batch.company,
                normalized=row.normalized_data_json,
                lock=True,
            )
            grouped_rows.setdefault(row_data["warehouse"].pk, []).append(
                (row, row_data)
            )
        for warehouse_id in sorted(grouped_rows, key=str):
            grouped = grouped_rows[warehouse_id]
            warehouse = grouped[0][1]["warehouse"]
            document = create_supply_document(
                actor=actor,
                company=batch.company,
                document_type="opening",
                data={
                    "business_date": timezone.localdate(),
                    "target_warehouse": warehouse,
                    "idempotency_key": (
                        f"opening-stock-import:{batch.pk}:{warehouse.pk}"
                    ),
                    "remark": f"由期初库存导入批次 {batch.pk} 生成，尚未过账。",
                },
                lines=[
                    {
                        "item": row_data["item"],
                        "quantity": row_data["quantity"],
                        "entered_unit_cost": row_data["entered_unit_cost"],
                        "line_remark": row_data["line_remark"],
                    }
                    for _, row_data in grouped
                ],
                request=request,
            )
            created_lines = list(document.lines.order_by("line_no"))
            if len(created_lines) != len(grouped):
                raise ValidationError("期初导入生成的单据明细数量不一致，已整批回滚。")
            for (row, _), document_line in zip(
                grouped, created_lines, strict=True
            ):
                row.created_object_type = "SupplyDocumentLine"
                row.created_object_id = str(document_line.pk)
    elif batch.import_type == "opening_custody":
        from apps.supplies.services import create_opening_custody_from_import_row

        for row in rows:
            row_data = _inflate_opening_custody_row(
                company=batch.company,
                normalized=row.normalized_data_json,
                lock=True,
            )
            custody = create_opening_custody_from_import_row(
                actor=actor,
                import_row=row,
                request=request,
                **row_data,
            )
            row.created_object_type = "SupplyCustody"
            row.created_object_id = str(custody.pk)
    elif batch.import_type == "department":
        existing = {
            value.normalized_code: value
            for value in Department.objects.select_for_update().filter(company=batch.company)
        }
        if any(row.normalized_data_json["normalized_code"] in existing for row in rows):
            raise ValidationError("确认前检测到部门编码已存在，请重新上传验证。")
        employees = {
            value.normalized_employee_no: value
            for value in Employee.objects.filter(company=batch.company).select_related("department")
        }
        pending = list(rows)
        while pending:
            progressed = False
            for row in pending[:]:
                item = row.normalized_data_json
                parent_code = item["normalized_parent_code"]
                parent = existing.get(parent_code) or created.get(parent_code)
                if parent_code and parent is None:
                    continue
                manager = employees.get(item["normalized_manager_employee_no"])
                department = create_department(
                    actor=actor,
                    company=batch.company,
                    data={
                        "code": item["code"],
                        "name": item["name"],
                        "parent": parent,
                        "manager_employee": manager,
                        "is_active": item["is_active"],
                    },
                    request=request,
                )
                created[item["normalized_code"]] = department
                row.created_object_type = "Department"
                row.created_object_id = str(department.pk)
                pending.remove(row)
                progressed = True
            if not progressed:
                raise ValidationError("无法解析部门树的父级顺序，已整批回滚。")
    elif batch.import_type == "employee":
        existing_numbers = set(
            Employee.objects.select_for_update().filter(company=batch.company).values_list(
                "normalized_employee_no", flat=True
            )
        )
        if any(row.normalized_data_json["normalized_employee_no"] in existing_numbers for row in rows):
            raise ValidationError("确认前检测到员工编号已存在，请重新上传验证。")
        departments = {
            value.normalized_code: value
            for value in Department.objects.filter(company=batch.company, is_active=True)
        }
        for row in rows:
            item = row.normalized_data_json
            department = departments.get(item["normalized_department_code"])
            if department is None:
                raise ValidationError(f"第 {row.row_number} 行部门已不可用，已整批回滚。")
            employee = create_employee(
                actor=actor,
                company=batch.company,
                data={
                    "employee_no": item["employee_no"],
                    "name": item["name"],
                    "department": department,
                    "employment_status": item["employment_status"],
                    "hire_date": date.fromisoformat(item["hire_date"]) if item["hire_date"] else None,
                    "termination_date": date.fromisoformat(item["termination_date"]) if item["termination_date"] else None,
                    "mobile": item["mobile"],
                    "remark": item["remark"],
                    "is_active": item["is_active"],
                },
                request=request,
            )
            row.created_object_type = "Employee"
            row.created_object_id = str(employee.pk)
    else:
        raise ValidationError("不支持的导入类型。")

    batch.status = "confirmed"
    batch.confirmed_by = actor
    batch.confirmed_at = timezone.now()
    batch.full_clean()
    batch.save(update_fields=["status", "confirmed_by", "confirmed_at"])
    if connection.vendor == "postgresql":
        with connection.cursor() as cursor:
            cursor.execute(
                "SELECT set_config('eam_lite.controlled_import_confirmation', 'on', true)"
            )
    for row in rows:
        row.validation_status = "created"
        row.full_clean()
        row.save(
            update_fields=[
                "validation_status",
                "created_object_type",
                "created_object_id",
            ]
        )
    if connection.vendor == "postgresql":
        with connection.cursor() as cursor:
            cursor.execute(
                "SELECT set_config('eam_lite.controlled_import_confirmation', 'off', true)"
            )
    _audit_batch(
        batch=batch,
        actor=actor,
        action="import_confirm",
        old_data={"status": "validated"},
        new_data={
            "status": "confirmed",
            "import_type": batch.import_type,
            "created_count": len(rows),
            "file_sha256": batch.file_sha256,
            "created_objects": [
                {
                    "row_number": row.row_number,
                    "object_type": row.created_object_type,
                    "object_id": row.created_object_id,
                }
                for row in rows
            ],
        },
        request=request,
    )
    return batch
